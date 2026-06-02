import logging
import datetime
import re
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
import io

from .models import (
    VendorMasterBasicDetail,
    VendorMasterGSTDetails,
    VendorMasterBanking,
    VendorMasterTDS,
    VendorMasterTerms,
    VendorMasterCategory
)
from .vendorproduct_database import VendorProductServiceDatabase
from accounting.models import MasterLedger

from core.utils import match_headers

logger = logging.getLogger(__name__)

# Column Definitions for Vendor Excel
VENDOR_COLUMNS = [
    {"label": "Vendor Name", "key": "vendor_name", "required": True},
    {"label": "Vendor Code", "key": "vendor_code", "required": True},
    {"label": "Category", "key": "category", "required": False},
    {"label": "PAN Number", "key": "pan_no", "required": False},
    {"label": "Contact Person", "key": "contact_person", "required": False},
    {"label": "Email Address", "key": "email", "required": True},
    {"label": "Contact Number", "key": "contact_no", "required": True},
    {"label": "Billing Currency", "key": "billing_currency", "required": False},
    {"label": "Registration Type", "key": "registration_type", "required": False},
    {"label": "GSTIN", "key": "gstin", "required": False},
    {"label": "Reference Name", "key": "reference_name", "required": False},
    {"label": "Address Line 1", "key": "branch_address_line1", "required": False},
    {"label": "Address Line 2", "key": "branch_address_line2", "required": False},
    {"label": "Address Line 3", "key": "branch_address_line3", "required": False},
    {"label": "City", "key": "branch_city", "required": True},
    {"label": "State", "key": "branch_state", "required": True},
    {"label": "Pincode", "key": "branch_pincode", "required": False},
    {"label": "Country", "key": "branch_country", "required": True},
    {"label": "Branch Contact Person", "key": "branch_contact_person", "required": False},
    {"label": "Branch Email Address", "key": "branch_email", "required": False},
    {"label": "Branch Contact Number", "key": "branch_contact_no", "required": False},
    {"label": "MSME No", "key": "msme_udyam_no", "required": False},
    {"label": "FSSAI No", "key": "fssai_license_no", "required": False},
    {"label": "IEC Code", "key": "import_export_code", "required": False},
    {"label": "TDS Section", "key": "tds_section", "required": False},
    {"label": "TCS Applicable", "key": "tcs_applicable", "required": False},
    {"label": "TCS Section", "key": "tcs_section", "required": False},
    {"label": "Credit Period", "key": "credit_period", "required": False},
    {"label": "Credit Limit", "key": "credit_limit", "required": False},
    {"label": "Credit Terms", "key": "credit_terms", "required": False},
    {"label": "Penalty Terms", "key": "penalty_terms", "required": False},
    {"label": "Delivery Terms", "key": "delivery_terms", "required": False},
    {"label": "Warranty Details", "key": "warranty_guarantee_details", "required": False},
    {"label": "Force Majeure", "key": "force_majeure", "required": False},
    {"label": "Dispute Terms", "key": "dispute_redressal_terms", "required": False},
    {"label": "Bank Account No", "key": "bank_account_no", "required": False},
    {"label": "Bank Name", "key": "bank_name", "required": False},
    {"label": "IFSC Code", "key": "ifsc_code", "required": False},
    {"label": "Bank Branch", "key": "branch_name", "required": False},
    {"label": "Swift Code", "key": "swift_code", "required": False},
    {"label": "Associated Branch", "key": "vendor_branch", "required": False},
    {"label": "Item Code", "key": "item_code", "required": False},
    {"label": "Item Name", "key": "item_name", "required": False},
    {"label": "HSN/SAC Code", "key": "hsn_code", "required": False},
    {"label": "Supplier Item Code", "key": "supplier_item_code", "required": False},
    {"label": "Supplier Item Name", "key": "supplier_item_name", "required": False},
    {"label": "Packing Notes", "key": "packing_notes", "required": False},
]

class VendorExcelTemplateDownloadView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        wb = Workbook()
        
        # 1. Create Readme & Instructions sheet as the active/first sheet
        ws_readme = wb.active
        assert ws_readme is not None
        ws_readme.title = "Instructions & Readme"
        
        # Title Block
        ws_readme.merge_cells("A1:D1")
        title_cell = ws_readme["A1"]
        title_cell.value = "Vendor Master Creation - Excel Upload Instructions"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_readme.row_dimensions[1].height = 40
        
        # Subtitle / Note
        ws_readme.merge_cells("A2:D2")
        sub_cell = ws_readme["A2"]
        sub_cell.value = "Please read these guidelines carefully before populating the 'Vendor Template' tab to ensure a successful upload."
        sub_cell.font = Font(size=10, italic=True, color="4B5563")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_readme.row_dimensions[2].height = 20
        
        # Table Headers
        headers = ["Field Label", "Mandatory / Optional", "Format & Length", "Description & Guidelines"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_readme.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(border_style="thin", color="D1D5DB")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws_readme.row_dimensions[4].height = 25
        
        # Instruction Rows grouped by Sections matching UI Tabs
        readme_data = [
            {"section": "1. BASIC DETAILS"},
            ("Vendor Name", "Mandatory", "Text", "Full legal entity name or trade name of the vendor."),
            ("Category", "Mandatory", "Text", "Must match an existing vendor category (e.g., Raw Material, Packing Material)."),
            ("Email Address", "Mandatory", "Email Format", "Primary communication email for sending purchase orders and notifications."),
            ("Contact Number", "Mandatory", "Numeric / Text", "Primary contact phone or mobile number."),
            ("Vendor Code", "Mandatory", "Text", "Custom unique vendor identifier. VEN-XXXXXX format."),
            ("PAN Number", "Optional", "10 Characters", "Must be exactly 10 alphanumeric characters in standard PAN format (e.g., ABCDE1234F)."),
            ("Billing Currency", "Optional", "Text (e.g., INR, USD)", "Default billing currency for purchase orders. Defaults to INR if empty."),
            
            {"section": "2. BRANCH DETAILS"},
            ("Reference Name", "Mandatory", "Text", "Name of the vendor's billing/shipping branch or location (e.g., Main Branch, Factory)."),
            ("Address Line 1", "Mandatory", "Text", "Flat/House no., building name, or street address."),
            ("Address Line 2", "Mandatory", "Text", "Locality, sector, area, or road name."),
            ("Registration Type", "Optional", "Text", "Registration type of the vendor (e.g., Regular, Composition). Defaults to Regular."),
            ("GSTIN", "Optional", "15 Characters", "Must be exactly 15 alphanumeric characters in valid GSTIN format starting with state code."),
            ("Address Line 3, City, State, Pincode, Country", "Optional", "Text / Numeric", "Additional detailed address and geographical identifiers."),
            
            {"section": "3. PRODUCTS/SERVICES"},
            ("Item Code, Item Name", "Optional", "Text", "Initial product or service mapping supplied by this vendor."),
            ("HSN/SAC Code", "Optional", "Numeric", "HSN/SAC classification code. Must be entirely numeric if provided."),
            ("Supplier Item Code / Name", "Optional", "Text", "Vendor's internal item code and description mapping."),
            
            {"section": "4. TDS & OTHER STATUTORY DETAILS"},
            ("TDS Section", "Optional", "Text", "Applicable Income Tax TDS section for bill deduction."),
            ("TCS Applicable / Section", "Optional", "Text", "Specify if GST/Income Tax TCS is applicable for this vendor."),
            ("MSME / FSSAI / IEC Numbers", "Optional", "Text / Numeric", "Statutory registration numbers for regulatory mapping."),
            
            {"section": "5. BANKING INFO"},
            ("Bank Account No", "Optional", "Numeric / Text", "Vendor's bank account number for digital transfers."),
            ("Bank Name, Bank Branch", "Optional", "Text", "Name of the bank and specific branch location."),
            ("IFSC Code, Swift Code", "Optional", "Text / Numeric", "IFSC routing code (11 characters) or international Swift code."),
            
            {"section": "6. TERMS & CONDITIONS"},
            ("Credit Period, Credit Limit", "Optional", "Numeric / Text", "Agreed payment credit window and spending ceiling."),
            ("Credit terms, Penalty terms", "Optional", "Text", "Standard financial credit and late payment clauses."),
            ("Delivery terms, Warranty details", "Optional", "Text", "FOB/CIF shipping terms and product guarantee rules."),
            ("Force Majeure, Dispute terms", "Optional", "Text", "Legal contingencies and dispute resolution jurisdiction.")
        ]
        
        current_row = 5
        thin_side = Side(border_style="thin", color="E5E7EB")
        border_all = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        
        for item in readme_data:
            if isinstance(item, dict) and "section" in item:
                ws_readme.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                cell = ws_readme.cell(row=current_row, column=1, value=item["section"])
                cell.font = Font(size=11, bold=True, color="065F46")
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                
                for c in range(1, 5):
                    ws_readme.cell(row=current_row, column=c).border = border_all
                ws_readme.row_dimensions[current_row].height = 25
                current_row += 1
            else:
                fill_color = "FFFFFF" if current_row % 2 != 0 else "F9FAFB"
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                for col_idx, val in enumerate(item, 1):
                    cell = ws_readme.cell(row=current_row, column=col_idx, value=val)
                    cell.fill = row_fill
                    cell.border = border_all
                    cell.font = Font(size=10, color="1F2937")
                    
                    if col_idx == 1:
                        cell.font = Font(size=10, bold=True, color="111827")
                        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    elif col_idx == 2:
                        is_mand = (val == "Mandatory")
                        cell.font = Font(size=10, bold=is_mand, color="B91C1C" if is_mand else "4B5563")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 3:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                ws_readme.row_dimensions[current_row].height = 22
                current_row += 1
            
        # Set specific column widths for Readme sheet
        ws_readme.column_dimensions['A'].width = 28
        ws_readme.column_dimensions['B'].width = 22
        ws_readme.column_dimensions['C'].width = 25
        ws_readme.column_dimensions['D'].width = 65
        
        # 2. Create the actual Vendor Template sheet
        ws = wb.create_sheet(title="Vendor Template")
        assert ws is not None
        
        # Headers
        for col_idx, col_def in enumerate(VENDOR_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["label"])
            cell.font = Font(bold=True, color="FFFFFF")
            
            bg_color = "10B981" if col_def.get("required") else "6B7280" # Emerald for vendors
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.column_dimensions[cell.column_letter].width = 20

        ws.freeze_panes = "A2"
        
        # Enable sheet protection to prevent header modification
        ws.protection.sheet = True
        
        # Unlock rows 2 to 1000 for data entry (up to 1000 records)
        for row in range(2, 1001):
            for col in range(1, len(VENDOR_COLUMNS) + 1):
                ws.cell(row=row, column=col).protection = Protection(locked=False)
                
        # Add Data Validation for Registration Type
        reg_col_letter = None
        for col_idx, col_def in enumerate(VENDOR_COLUMNS, 1):
            if col_def["key"] == "registration_type":
                from openpyxl.utils import get_column_letter
                reg_col_letter = get_column_letter(col_idx)
                break
                
        if reg_col_letter:
            dv = DataValidation(type="list", formula1='"Regular,Composition,Special Economic Zone (SEZ),Unregistered"', allow_blank=True, showErrorMessage=True)
            dv.error = 'Your entry is not in the list. Please select from the dropdown options.'
            dv.errorTitle = 'Invalid Entry'
            dv.prompt = 'Please select from the list'
            dv.promptTitle = 'Registration Type'
            ws.add_data_validation(dv)
            dv.add(f'{reg_col_letter}2:{reg_col_letter}1000')
        
        wb.active = 0
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="vendor_import_template.xlsx"'
        wb.save(response)
        return response

class VendorExcelExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        tenant_id = getattr(request.user, "tenant_id", None)
        vendors = VendorMasterBasicDetail.objects.filter(tenant_id=tenant_id, is_active=True)
        
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Vendors"
        
        # Headers
        for col_idx, col_def in enumerate(VENDOR_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["label"])
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = 20

        # Data
        for row_idx, vendor in enumerate(vendors, 2):
            gst = vendor.gst_details.first()
            banking = vendor.banking_details.first()
            tds = vendor.tds_details.first()
            terms = vendor.terms_conditions.first()
            
            data = {
                "Vendor Name": vendor.vendor_name,
                "Vendor Code": vendor.vendor_code,
                "Category": vendor.vendor_category,
                "PAN Number": vendor.pan_no,
                "Contact Person": vendor.contact_person,
                "Email Address": vendor.email,
                "Contact Number": vendor.contact_no,
                "Billing Currency": vendor.billing_currency,
                "Registration Type": gst.gst_registration_type.title() if gst and gst.gst_registration_type else "Regular",
                "GSTIN": gst.gstin if gst else "",
                "Reference Name": gst.reference_name if gst else "",
                "Address Line 1": gst.branch_address_line1 if gst else "",
                "Address Line 2": gst.branch_address_line2 if gst else "",
                "Address Line 3": gst.branch_address_line3 if gst else "",
                "City": gst.branch_city if gst else "",
                "State": gst.branch_state if gst else "",
                "Pincode": gst.branch_pincode if gst else "",
                "Country": gst.branch_country if (gst and gst.branch_country) else "India",
                "Branch Contact Person": gst.branch_contact_person if gst else "",
                "Branch Email Address": gst.branch_email if gst else "",
                "Branch Contact Number": gst.branch_contact_no if gst else "",
                "MSME No": tds.msme_udyam_no if tds else "",
                "FSSAI No": tds.fssai_license_no if tds else "",
                "IEC Code": tds.import_export_code if tds else "",
                "TDS Section": tds.tds_section_applicable if tds else "",
                "TCS Applicable": "Yes" if vendor.tcs_applicable else "No",
                "TCS Section": tds.tcs_section_applicable if tds else "",
                "Credit Period": terms.credit_period if terms else "",
                "Credit Limit": terms.credit_limit if terms else "",
                "Credit Terms": terms.credit_terms if terms else "",
                "Penalty Terms": terms.penalty_terms if terms else "",
                "Delivery Terms": terms.delivery_terms if terms else "",
                "Warranty Details": terms.warranty_guarantee_details if terms else "",
                "Force Majeure": terms.force_majeure if terms else "",
                "Dispute Terms": terms.dispute_redressal_terms if terms else "",
                "Bank Account No": banking.bank_account_no if banking else "",
                "Bank Name": banking.bank_name if banking else "",
                "IFSC Code": banking.ifsc_code if banking else "",
                "Bank Branch": banking.branch_name if banking else "",
            }
            
            # Fetch products
            prod_data = VendorProductServiceDatabase.get_by_vendor(vendor.id)
            if prod_data and prod_data.get("items"):
                p = prod_data["items"][0]
                data.update({
                    "Item Code": p.get("item_code"),
                    "Item Name": p.get("item_name"),
                    "HSN/SAC Code": p.get("hsn_sac_code"),
                    "Supplier Item Code": p.get("supplier_item_code"),
                    "Supplier Item Name": p.get("supplier_item_name"),
                })
            
            for col_idx, col_def in enumerate(VENDOR_COLUMNS, 1):
                ws.cell(row=row_idx, column=col_idx, value=data.get(str(col_def["label"]), ""))

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="vendors_export.xlsx"'
        wb.save(response)
        return response

class VendorExcelUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        tenant_id = getattr(request.user, "tenant_id", None)
        username = getattr(request.user, "username", "system")
        dry_run = request.query_params.get("dry_run") == "true"
        
        excel_file = request.FILES.get("file")
        json_data = request.data.get("data")

        if not excel_file and not json_data:
            return Response({"error": "No file or JSON data provided"}, status=400)

        # DEBUG: Save file to disk
        if excel_file:
            try:
                import os
                from django.conf import settings
                with open(os.path.join(settings.BASE_DIR, 'debug_excel.xlsx'), 'wb') as f:
                    for chunk in excel_file.chunks():
                        f.write(chunk)
                excel_file.seek(0)
            except Exception as e:
                print("Failed to save debug excel", e)
        
        try:
            records_to_process = []
            
            if excel_file:
                wb = load_workbook(excel_file, data_only=True)
                ws = wb.active
                assert ws is not None
                
                # Header mapping using fuzzy matching
                excel_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
                header_index = match_headers(excel_headers, VENDOR_COLUMNS)
                
                # Validate required columns
                for col in VENDOR_COLUMNS:
                    if col["required"] and col["label"] not in header_index:
                        return Response({"error": f"Missing required column: {col['label']}"}, status=400)

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if not any(row): continue
                    row_data = {}
                    for lbl, idx in header_index.items():
                        row_data[lbl] = row[idx-1]
                    
                    # Normalize: fix Reference Name if it got wrongly mapped to Bank Branch key due to header conflict
                    # This happens because 'branch_name' normalizes to 'branchname', matching 'Reference Name'
                    def _raw_empty(v):
                        return not v or str(v).strip().lower() in ['', 'none', 'nan', 'null', 'n/a']
                    
                    if _raw_empty(row_data.get("Reference Name")):
                        for _fallback in ["Bank Branch", "reference_name", "BranchName"]:
                            if not _raw_empty(row_data.get(_fallback)):
                                row_data["Reference Name"] = row_data[_fallback]
                                break
                    
                    records_to_process.append({"row_data": row_data, "row_index": row_idx})
            else:
                if isinstance(json_data, str):
                    import json
                    records_to_process = json.loads(json_data)
                else:
                    records_to_process = json_data

            results = {"success": 0, "failed": 0, "errors": [], "successful_imports": []}
            
            # Group hierarchical rows (subsequent rows with empty vendor basic details)
            grouped_records = []
            current_main_record = None

            def _raw_empty(v):
                return not v or str(v).strip().lower() in ['', 'none', 'nan', 'null', 'n/a']

            for item in records_to_process:
                r_data = item.get("row_data")
                if not r_data: continue

                v_code = r_data.get("Vendor Code") or r_data.get("vendor_code")
                v_name = r_data.get("Vendor Name") or r_data.get("vendor_name")
                email = r_data.get("Email Address") or r_data.get("email")

                is_main = not (_raw_empty(v_code) and _raw_empty(v_name) and _raw_empty(email))

                if is_main:
                    if "extra_branches" not in r_data: r_data["extra_branches"] = []
                    if "extra_banks" not in r_data: r_data["extra_banks"] = []
                    if "products" not in r_data: r_data["products"] = []
                    
                    has_product = not _raw_empty(r_data.get("Item Code") or r_data.get("item_code") or r_data.get("Item Name") or r_data.get("item_name"))
                    if has_product:
                        # Extract product data from the main row to include in the products list
                        r_data["products"].append({
                            "Item Code": r_data.get("Item Code"),
                            "Item Name": r_data.get("Item Name"),
                            "HSN/SAC Code": r_data.get("HSN/SAC Code") or r_data.get("HSN/SAC") or r_data.get("hsn_code") or r_data.get("hsn_sac_code"),
                            "UOM": r_data.get("UOM"),
                            "Supplier Item Code": r_data.get("Supplier Item Code") or r_data.get("Supp Item Code"),
                            "Supplier Item Name": r_data.get("Supplier Item Name") or r_data.get("Supp Item Name"),
                            "Packing Notes": r_data.get("Packing Notes"),
                        })
                        
                    current_main_record = item
                    grouped_records.append(item)
                else:
                    if current_main_record:
                        # Append to current main record
                        has_branch = not _raw_empty(r_data.get("Reference Name") or r_data.get("branch_name") or r_data.get("reference_name") or r_data.get("City"))
                        has_bank = not _raw_empty(r_data.get("Bank Account No") or r_data.get("bank_account_no") or r_data.get("Bank Name") or r_data.get("bank_name"))
                        has_product = not _raw_empty(r_data.get("Item Code") or r_data.get("item_code") or r_data.get("Item Name") or r_data.get("item_name"))

                        if has_branch: current_main_record["row_data"]["extra_branches"].append(r_data)
                        if has_bank: current_main_record["row_data"]["extra_banks"].append(r_data)
                        if has_product: current_main_record["row_data"]["products"].append(r_data)
                    else:
                        # No preceding main record, treat as main so it fails validation gracefully
                        grouped_records.append(item)
            
            records_to_process = grouped_records
            
            # Cache for vendors created in this file (so multiple distinct named branches in same file link together)
            file_vendors_cache = {}
            file_vendors_code_cache = {}

            for item in records_to_process:
                row_data = item.get("row_data")
                row_idx = item.get("row_index", "N/A")
                
                if not row_data: continue
                
                v_code = row_data.get("Vendor Code") or row_data.get("vendor_code")
                v_name = row_data.get("Vendor Name") or row_data.get("vendor_name")
                email = row_data.get("Email Address") or row_data.get("email")
                contact = row_data.get("Contact Number") or row_data.get("contact_no")
                
                def is_empty(val):
                    if not val: return True
                    s = str(val).strip().lower()
                    return s in ['n/a', 'none', 'nan', 'null', ''] or s.startswith('select ')

                # === Comprehensive mandatory field check ===
                # Collect ALL missing mandatory fields at once
                mandatory_checks = [
                    ("Vendor Code",    row_data.get("Vendor Code") or row_data.get("vendor_code")),
                    ("Vendor Name",    row_data.get("Vendor Name") or row_data.get("vendor_name")),
                    ("Email Address",  row_data.get("Email Address") or row_data.get("email")),
                    ("Contact Number", row_data.get("Contact Number") or row_data.get("contact_no")),
                    ("Category",       row_data.get("Category") or row_data.get("vendor_category")),
                    ("PAN Number",     row_data.get("PAN Number") or row_data.get("pan_no")),
                    ("Reference Name", (row_data.get("Reference Name") or row_data.get("branch_name")
                                        or row_data.get("reference_name") or row_data.get("Bank Branch"))),
                    ("Address Line 1", row_data.get("Address Line 1") or row_data.get("address_line_1")),
                    ("Address Line 2", row_data.get("Address Line 2") or row_data.get("address_line_2")),
                    ("Country",        row_data.get("Country") or row_data.get("branch_country")),
                    ("State",          row_data.get("State") or row_data.get("branch_state")),
                    ("City",           row_data.get("City") or row_data.get("branch_city")),
                ]
                missing_fields = [name for name, val in mandatory_checks if is_empty(val)]
                if missing_fields:
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: {', '.join(missing_fields)} is missing",
                        "missing_fields": missing_fields,
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue
                
                # Validation
                pan = row_data.get("PAN Number") or row_data.get("pan_no")
                if pan:
                    if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]{1}$', str(pan).strip().upper()):
                        results["failed"] += 1
                        results["errors"].append({
                            "message": f"Row {row_idx}: Invalid PAN format (must be AAAAA0000A)",
                            "row_data": row_data,
                            "row_index": row_idx
                        })
                        continue
                
                gstin = row_data.get("GSTIN") or row_data.get("gstin")
                if gstin:
                    if not re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$', str(gstin).strip().upper()):
                        results["failed"] += 1
                        results["errors"].append({
                            "message": f"Row {row_idx}: Invalid GSTIN format",
                            "row_data": row_data,
                            "row_index": row_idx
                        })
                        continue
                
                try:
                    # Use a nested transaction that we can rollback if dry_run
                    with transaction.atomic():
                        # Check DB for duplicate / soft-deleted vendor
                        vendor = None
                        v_name_lower = str(v_name).strip().lower()
                        
                        # Check cache first for same-file duplicates
                        if v_code and v_code in file_vendors_code_cache:
                            raise Exception(f"DUPLICATE ENTRY: Vendor Code '{v_code}' appears more than once in this file. Each vendor must have a unique code.")
                        
                        if not vendor:
                            # Search by vendor code
                            if v_code:
                                existing_by_code = VendorMasterBasicDetail.objects.filter(tenant_id=tenant_id, vendor_code=v_code).first()
                                if existing_by_code:
                                    if not existing_by_code.is_deleted:
                                        raise Exception(f"DUPLICATE ENTRY: Vendor Code '{v_code}' already exists in the system. Please use a unique code or leave it blank.")
                                    else:
                                        # It's soft-deleted, we will reactivate it
                                        vendor = existing_by_code

                        if vendor:
                            # Restore/reactivate soft-deleted vendor
                            vendor.is_deleted = False
                            vendor.is_active = True
                            vendor.vendor_name = v_name
                            vendor.vendor_code = v_code or vendor.vendor_code
                            vendor.pan_no = pan
                            vendor.contact_person = row_data.get("Contact Person") or row_data.get("contact_person")
                            vendor.email = email
                            vendor.contact_no = str(contact)
                            vendor.vendor_category = row_data.get("Category") or row_data.get("category")
                            vendor.billing_currency = row_data.get("Billing Currency") or row_data.get("billing_currency") or "INR"
                            vendor.tcs_applicable = True if str(row_data.get("TCS Applicable") or row_data.get("tcs_applicable", "")).lower() in ['yes', 'true', '1'] else False
                            vendor.updated_by = username
                            vendor.save()
                            
                            # Clean up existing related records to prevent duplicate constraints or orphan records
                            VendorMasterGSTDetails.objects.filter(vendor_basic_detail=vendor).delete()
                            VendorMasterBanking.objects.filter(vendor_basic_detail=vendor).delete()
                            VendorMasterTDS.objects.filter(vendor_basic_detail=vendor).delete()
                            VendorMasterTerms.objects.filter(vendor_basic_detail=vendor).delete()
                        else:
                            # Create new vendor
                            vendor = VendorMasterBasicDetail(
                                tenant_id=tenant_id,
                                vendor_name=v_name,
                                vendor_code=v_code,
                                pan_no=pan,
                                contact_person=row_data.get("Contact Person") or row_data.get("contact_person"),
                                email=email,
                                contact_no=str(contact),
                                vendor_category=row_data.get("Category") or row_data.get("category"),
                                billing_currency=row_data.get("Billing Currency") or row_data.get("billing_currency") or "INR",
                                tcs_applicable=True if str(row_data.get("TCS Applicable") or row_data.get("tcs_applicable", "")).lower() in ['yes', 'true', '1'] else False,
                                created_by=username
                            )
                            vendor.generate_vendor_code()
                            vendor.save()

                        # Populate same-file cache
                        file_vendors_cache[v_name_lower] = vendor
                        file_vendors_code_cache[vendor.vendor_code] = vendor
                        
                        # Ledger creation
                        from accounting.utils_ledger import get_or_create_entity_ledger
                        ledger = get_or_create_entity_ledger(
                            tenant_id=tenant_id,
                            entity_name=v_name,
                            entity_type='vendor',
                            created_by=username
                        )
                        vendor.ledger = ledger
                        vendor.save(update_fields=['ledger_id'])
                        
                        # 2. GST Details
                        all_branches = [row_data] + row_data.get("extra_branches", [])
                        for branch_data in all_branches:
                            gstin = branch_data.get("GSTIN") or branch_data.get("gstin") or ""
                            if gstin or branch_data.get("Reference Name") or branch_data.get("reference_name"):
                                    
                                    raw_reg = str(branch_data.get("Registration Type") or branch_data.get("registration_type") or "regular").strip().lower()
                                    reg_map = {
                                        "regular": "regular",
                                        "composition": "composition",
                                        "special economic zone (sez)": "special_economic_zone",
                                        "special economic zone": "special_economic_zone",
                                        "unregistered": "unregistered",
                                        "consumer": "consumer",
                                        "overseas": "overseas",
                                        "deemed export": "deemed_export"
                                    }
                                    
                                    # Ensure reference_name is unique to prevent unique_together violations on empty/same gstin
                                    ref_name = (branch_data.get("Reference Name") or branch_data.get("reference_name", "Main Branch")).strip()
                                    gstin_cleaned = str(gstin).strip()
                                    if VendorMasterGSTDetails.objects.filter(tenant_id=tenant_id, gstin=gstin_cleaned, reference_name__iexact=ref_name).exists():
                                        ref_name = f"{ref_name} - {vendor.vendor_code}"
                                    
                                    VendorMasterGSTDetails.objects.create(  # type: ignore
                                        tenant_id=tenant_id,
                                        vendor_basic_detail=vendor,
                                        gstin=gstin,
                                        gst_registration_type=reg_map.get(raw_reg, "regular"),
                                    legal_name=vendor.vendor_name, # Default to vendor name if not provided
                                    reference_name=ref_name,
                                    branch_address_line1=branch_data.get("Address Line 1") or branch_data.get("branch_address_line1"),
                                    branch_address_line2=branch_data.get("Address Line 2") or branch_data.get("branch_address_line2"),
                                    branch_address_line3=branch_data.get("Address Line 3") or branch_data.get("branch_address_line3"),
                                    branch_city=branch_data.get("City") or branch_data.get("branch_city"),
                                    branch_state=branch_data.get("State") or branch_data.get("branch_state"),
                                    branch_pincode=branch_data.get("Pincode") or branch_data.get("branch_pincode"),
                                    branch_country=branch_data.get("Country") or branch_data.get("branch_country") or "India",
                                    branch_contact_person=branch_data.get("Branch Contact Person") or branch_data.get("branch_contact_person"),
                                    branch_email=branch_data.get("Branch Email Address") or branch_data.get("branch_email"),
                                    branch_contact_no=branch_data.get("Branch Contact Number") or branch_data.get("branch_contact_no"),
                                    created_by=username
                                )
                        
                        # 3. Banking Details
                        all_banks = [row_data] + row_data.get("extra_banks", [])
                        for bank_data in all_banks:
                            acc_no = bank_data.get("Bank Account No") or bank_data.get("bank_account_no")
                            if acc_no:
                                VendorMasterBanking.objects.create(  # type: ignore
                                    tenant_id=tenant_id,
                                    vendor_basic_detail=vendor,
                                    bank_account_no=str(acc_no),
                                    bank_name=bank_data.get("Bank Name") or bank_data.get("bank_name") or "",
                                    ifsc_code=bank_data.get("IFSC Code") or bank_data.get("ifsc_code") or "",
                                    branch_name=bank_data.get("Bank Branch") or bank_data.get("branch_name") or "",
                                    swift_code=bank_data.get("Swift Code") or bank_data.get("swift_code") or "",
                                    vendor_branch=bank_data.get("Associated Branch") or bank_data.get("vendor_branch") or "",
                                    created_by=username
                                )
                            elif bank_data is row_data:
                                # Create empty one only for the primary row if missing
                                VendorMasterBanking.objects.create(tenant_id=tenant_id, vendor_basic_detail=vendor, created_by=username)  # type: ignore
 
                        # 4. TDS Details
                        VendorMasterTDS.objects.create(  # type: ignore
                            tenant_id=tenant_id,
                            vendor_basic_detail=vendor,
                            msme_udyam_no=row_data.get("MSME No") or row_data.get("msme_udyam_no"),
                            fssai_license_no=row_data.get("FSSAI No") or row_data.get("fssai_license_no"),
                            import_export_code=row_data.get("IEC Code") or row_data.get("import_export_code"),
                            tds_section_applicable=row_data.get("TDS Section") or row_data.get("tds_section"),
                            tcs_section_applicable=row_data.get("TCS Section") or row_data.get("tcs_section"),
                            created_by=username
                        )
                        
                        # 5. Terms and Conditions
                        VendorMasterTerms.objects.create(  # type: ignore
                            tenant_id=tenant_id,
                            vendor_basic_detail=vendor,
                            credit_period=row_data.get("Credit Period") or row_data.get("credit_period"),
                            credit_limit=row_data.get("Credit Limit") or row_data.get("credit_limit"),
                            credit_terms=row_data.get("Credit Terms") or row_data.get("credit_terms"),
                            penalty_terms=row_data.get("Penalty Terms") or row_data.get("penalty_terms"),
                            delivery_terms=row_data.get("Delivery Terms") or row_data.get("delivery_terms"),
                            warranty_guarantee_details=row_data.get("Warranty Details") or row_data.get("warranty_guarantee_details"),
                            force_majeure=row_data.get("Force Majeure") or row_data.get("force_majeure"),
                            dispute_redressal_terms=row_data.get("Dispute Terms") or row_data.get("dispute_redressal_terms"),
                            created_by=username
                        )
                        
                        # 6. Products/Services
                        products_raw = row_data.get("products", [])
                        products_to_save = []
                        if products_raw:
                            for p in products_raw:
                                products_to_save.append({
                                    "item_code": p.get("Item Code") or p.get("item_code") or "",
                                    "item_name": p.get("Item Name") or p.get("item_name") or "",
                                    "hsn_sac_code": p.get("HSN/SAC Code") or p.get("HSN/SAC") or p.get("hsn_sac_code") or "",
                                    "uom": p.get("UOM") or p.get("uom") or "",
                                    "supplier_item_code": p.get("Supplier Item Code") or p.get("Supp Item Code") or p.get("supplier_item_code") or "",
                                    "supplier_item_name": p.get("Supplier Item Name") or p.get("Supp Item Name") or p.get("supplier_item_name") or "",
                                    "packing_notes": p.get("Packing Notes") or p.get("packing_notes") or "",
                                })
                        else:
                            # Try to get from flat fields
                            item_code = row_data.get("Item Code") or row_data.get("item_code")
                            if item_code:
                                products_to_save = [{
                                    "item_code": item_code,
                                    "item_name": row_data.get("Item Name") or row_data.get("item_name"),
                                    "hsn_sac_code": row_data.get("HSN/SAC Code") or row_data.get("HSN/SAC") or row_data.get("hsn_code") or row_data.get("hsn_sac_code"),
                                    "uom": row_data.get("UOM") or row_data.get("uom"),
                                    "supplier_item_code": row_data.get("Supplier Item Code") or row_data.get("Supp Item Code") or row_data.get("supplier_item_code"),
                                    "supplier_item_name": row_data.get("Supplier Item Name") or row_data.get("Supp Item Name") or row_data.get("supplier_item_name"),
                                    "packing_notes": row_data.get("Packing Notes") or row_data.get("packing_notes"),
                                }]
                        
                        if products_to_save:
                            # HSN Validation
                            for p in products_to_save:
                                hsn = p.get("hsn_sac_code")
                                if hsn and not str(hsn).replace(" ", "").isdigit():
                                    raise Exception(f"Invalid HSN/SAC Code: '{hsn}'. HSN codes must be numeric.")

                            VendorProductServiceDatabase.upsert_product_services(
                                tenant_id=tenant_id,
                                vendor_basic_detail_id=getattr(vendor, 'id', None),
                                items=products_to_save,
                                created_by=username
                            )
                        
                        results["successful_imports"].append({
                            "id": None if dry_run else getattr(vendor, 'id', None),
                            "name": vendor.vendor_name,
                            "code": vendor.vendor_code,
                            "row_data": row_data
                        })
                        results["success"] += 1

                        if dry_run:
                            # Rollback this specific row
                            raise Exception("Dry Run Rollback")
                        
                except Exception as row_err:
                    if str(row_err) == "Dry Run Rollback":
                        continue # This is expected for dry run, we already counted success
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: {str(row_err)}",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    logger.error(f"Error processing vendor row {row_idx}: {row_err}")

            return Response({
                "message": "Preview complete" if dry_run else f"Processing complete. Success: {results['success']}, Failed: {results['failed']}",
                "summary": results,
                "is_preview": dry_run
            }, status=200)

        except Exception as e:
            logger.error(f"Vendor Excel upload failed: {e}", exc_info=True)
            return Response({"error": f"Internal error: {str(e)}"}, status=500)
