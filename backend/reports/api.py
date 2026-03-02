"""
Reports API Layer
"""
from rest_framework import views
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.http import HttpResponse
import openpyxl
from datetime import datetime

class PlaceholderReportView(views.APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        return Response({"message": "Reports module under construction"})


class DaybookExcelView(views.APIView):
    """Export daybook report as Excel file"""
    permission_classes = [AllowAny]  # Temporary for development
    
    def get(self, request):
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daybook"
        
        # Add headers
        headers = ['Date', 'Voucher No', 'Voucher Type', 'Party', 'Debit', 'Credit', 'Narration']
        ws.append(headers)
        
        # Add sample data
        sample_data = [
            [datetime.now().strftime('%Y-%m-%d'), 'SALES-001', 'Sales', 'ABC Ltd', 10000, 0, 'Sales invoice'],
            [datetime.now().strftime('%Y-%m-%d'), 'PURCH-001', 'Purchase', 'XYZ Suppliers', 0, 5000, 'Purchase invoice'],
        ]
        
        for row in sample_data:
            ws.append(row)
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=daybook_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response


class TrialBalanceExcelView(views.APIView):
    """Export trial balance report as Excel file"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trial Balance"
        
        headers = ['Account Name', 'Debit', 'Credit']
        ws.append(headers)
        
        sample_data = [
            ['Cash', 50000, 0],
            ['Bank', 100000, 0],
            ['Sales', 0, 150000],
            ['Purchases', 80000, 0],
        ]
        
        for row in sample_data:
            ws.append(row)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=trial_balance_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response


class StockSummaryExcelView(views.APIView):
    """Export stock summary report as Excel file"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Summary"
        
        headers = ['Item Name', 'Opening Stock', 'Inward', 'Outward', 'Closing Stock']
        ws.append(headers)
        
        sample_data = [
            ['Product A', 100, 50, 30, 120],
            ['Product B', 200, 100, 80, 220],
        ]
        
        for row in sample_data:
            ws.append(row)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=stock_summary_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response


class LedgerExcelView(views.APIView):
    """Export ledger report as Excel file"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ledger Report"
        
        headers = ['Date', 'Particulars', 'Debit', 'Credit', 'Balance']
        ws.append(headers)
        
        sample_data = [
            [datetime.now().strftime('%Y-%m-%d'), 'Opening Balance', 0, 0, 50000],
            [datetime.now().strftime('%Y-%m-%d'), 'Sales', 10000, 0, 60000],
        ]
        
        for row in sample_data:
            ws.append(row)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=ledger_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response


class GSTExcelView(views.APIView):
    """Export GST report as Excel file"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GST Report"
        
        headers = ['GSTIN', 'Party Name', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total']
        ws.append(headers)
        
        sample_data = [
            ['22AAAAA0000A1Z5', 'ABC Ltd', 50000, 4500, 4500, 0, 59000],
        ]
        
        for row in sample_data:
            ws.append(row)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=gst_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response

class AIReportExcelView(views.APIView):
    """Export AI report as Excel file"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AI Report"
        
        headers = ['Metric', 'Value', 'Analysis']
        ws.append(headers)
        
        # Sample data - in a real app this would come from the AI analysis
        sample_data = [
            ['Total Revenue', '₹ 1,50,000', 'Positive growth trend'],
            ['Top Customer', 'ABC Corp', 'Consistent purchaser'],
            ['Expense Trend', 'Stable', 'Expenses within budget'],
        ]
        
        for row in sample_data:
            ws.append(row)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=ai_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response
