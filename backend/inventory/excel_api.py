import logging
import datetime
import re
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
import io

from .models import (
    InventoryItem,
    InventoryMasterCategory,
    InventoryUnit
)
from core.tenant import get_tenant_from_request
from core.utils import match_headers

logger = logging.getLogger(__name__)

# Column Definitions for Inventory Item Excel
INVENTORY_ITEM_COLUMNS = [
    {"label": "Item Code", "key": "item_code", "required": True},
    {"label": "Item Name", "key": "item_name", "required": True},
    {"label": "Description", "key": "description", "required": False},
    {"label": "Category Path", "key": "category_path", "required": False},
    {"label": "UOM", "key": "uom", "required": False},
    {"label": "Alternate UOM", "key": "alternate_uom", "required": False},
    {"label": "Conversion Factor", "key": "conversion_factor", "required": False},
    {"label": "Rate", "key": "rate", "required": False},
    {"label": "Rate Unit", "key": "rate_unit", "required": False},
    {"label": "HSN Code", "key": "hsn_code", "required": False},
    {"label": "GST Rate", "key": "gst_rate", "required": False},
    {"label": "Cess Rate", "key": "cess_rate", "required": False},
    {"label": "Reorder Level", "key": "reorder_level", "required": False},
    {"label": "Is Saleable", "key": "is_saleable", "required": False},
]

class InventoryItemExcelTemplateDownloadView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        wb = Workbook()
        
        # 1. Create Instructions sheet
        ws_readme = wb.active
        assert ws_readme is not None
        ws_readme.title = "Instructions & Readme"
        
        # Title Block
        ws_readme.merge_cells("A1:D1")
        title_cell = ws_readme["A1"]
        title_cell.value = "Inventory Item Creation - Excel Upload Instructions"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo theme
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_readme.row_dimensions[1].height = 40
        
        # Subtitle / Note
        ws_readme.merge_cells("A2:D2")
        sub_cell = ws_readme["A2"]
        sub_cell.value = "Please read these guidelines carefully before populating the 'Item Template' tab to ensure a successful upload."
        sub_cell.font = Font(size=10, italic=True, color="4B5563")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_readme.row_dimensions[2].height = 20
        
        # Table Headers
        headers = ["Field Label", "Mandatory / Optional", "Format & Length", "Description & Guidelines"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_readme.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(border_style="thin", color="D1D5DB")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws_readme.row_dimensions[4].height = 25
        
        readme_data = [
            {"section": "INVENTORY ITEM DETAILS"},
            ("Item Code", "Mandatory", "Text", "Unique identifier for the item (e.g., ITEM001, RM-002)."),
            ("Item Name", "Mandatory", "Text", "Descriptive name of the inventory item."),
            ("Description", "Optional", "Text", "Detailed description of the item."),
            ("Category Path", "Optional", "Text", "Hierarchy separated by '>' (e.g., Raw Material > Import > HI). Automatically creates category steps if not found."),
            ("UOM", "Optional", "Text", "Unit of Measure (e.g., nos, kg, mtr, ltr). Defaults to 'nos' if empty."),
            ("Alternate UOM", "Optional", "Text", "Alternate unit of measure for conversion."),
            ("Conversion Factor", "Optional", "Numeric", "Conversion factor between main and alternate UOM."),
            ("Rate", "Optional", "Numeric", "Standard cost or price rate of the item. Defaults to 0.00."),
            ("Rate Unit", "Optional", "Text", "Unit to which the rate applies. Defaults to UOM if empty."),
            ("HSN Code", "Optional", "Numeric / Text", "HSN or SAC classification code."),
            ("GST Rate", "Optional", "Numeric / Percent", "Tax rate percentage (e.g. 18, 18%, 5, 0)."),
            ("Cess Rate", "Optional", "Numeric", "Compensation cess rate percentage (e.g. 1.5, 0)."),
            ("Reorder Level", "Optional", "Numeric", "Threshold quantity to trigger reorder alert."),
            ("Is Saleable", "Optional", "Yes / No", "Specify 'Yes' or 'No' if this item is for sales. Defaults to 'No'.")
        ]
        
        current_row = 5
        thin_side = Side(border_style="thin", color="E5E7EB")
        border_all = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        
        for item in readme_data:
            if isinstance(item, dict) and "section" in item:
                ws_readme.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                cell = ws_readme.cell(row=current_row, column=1, value=item["section"])
                cell.font = Font(size=11, bold=True, color="3730A3")
                cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                
                for c in range(1, 5):
                    ws_readme.cell(row=current_row, column=c).border = border_all
                ws_readme.row_dimensions[current_row].height = 25
                current_row += 1
            else:
                fill_color = "FFFFFF" if current_row % 2 != 0 else "F9FAFB"
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                for col_idx, val in enumerate(item, 1):
                    cell = ws_readme.cell(row=current_row, column=col_idx, value=val)
                    cell.fill = row_fill
                    cell.border = border_all
                    cell.font = Font(size=10, color="1F2937")
                    
                    if col_idx == 1:
                        cell.font = Font(size=10, bold=True, color="111827")
                        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    elif col_idx == 2:
                        is_mand = (val == "Mandatory")
                        cell.font = Font(size=10, bold=is_mand, color="B91C1C" if is_mand else "4B5563")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 3:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                ws_readme.row_dimensions[current_row].height = 22
                current_row += 1
            
        ws_readme.column_dimensions['A'].width = 28
        ws_readme.column_dimensions['B'].width = 22
        ws_readme.column_dimensions['C'].width = 25
        ws_readme.column_dimensions['D'].width = 65
        
        # 2. Create the Item Template sheet
        ws = wb.create_sheet(title="Item Template")
        assert ws is not None
        
        # Headers
        for col_idx, col_def in enumerate(INVENTORY_ITEM_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["label"])
            cell.font = Font(bold=True, color="FFFFFF")
            
            bg_color = "4F46E5" if col_def.get("required") else "6B7280" # Indigo for required, Gray for optional
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.column_dimensions[cell.column_letter].width = 20

        ws.freeze_panes = "A2"
        
        # Enable sheet protection
        ws.protection.sheet = True
        
        # Unlock rows 2 to 1000 for data entry
        for row in range(2, 1001):
            for col in range(1, len(INVENTORY_ITEM_COLUMNS) + 1):
                ws.cell(row=row, column=col).protection = Protection(locked=False)
        
        wb.active = 0
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="inventory_item_import_template.xlsx"'
        wb.save(response)
        return response

class InventoryItemExcelUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        tenant_id = get_tenant_from_request(request)
        username = getattr(request.user, "username", "system")
        excel_file = request.FILES.get("file")
        dry_run = request.query_params.get("dry_run") == "true"

        # --- PATH A: CONFIRM IMPORT — frontend sends JSON array of pre-validated rows ---
        raw_data = request.data.get("data")
        if not excel_file and raw_data:
            import json as _json
            try:
                records_to_process = _json.loads(raw_data) if isinstance(raw_data, str) else raw_data
            except Exception:
                return Response({"error": "Invalid JSON data"}, status=400)

            results = {"success": 0, "failed": 0, "errors": [], "successful_imports": []}
            seen_item_codes = set()

            def is_empty(val):
                return not val or str(val).strip().lower() in ['n/a', 'none', 'nan', 'null', '']

            for item in records_to_process:
                row_data = item.get("row_data", {})
                row_idx = item.get("row_index", "N/A")
                if not row_data:
                    continue

                item_code = row_data.get("Item Code")
                item_name = row_data.get("Item Name")

                if is_empty(item_code) or is_empty(item_name):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Item Code or Item Name is missing",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                item_code_clean = str(item_code).strip()
                if item_code_clean in seen_item_codes:
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Duplicate Item Code '{item_code_clean}' in the import list",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue
                seen_item_codes.add(item_code_clean)

                if InventoryItem.objects.filter(tenant_id=tenant_id, item_code=item_code_clean, is_active=True).exists():
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Item Code '{item_code_clean}' already exists in the database",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                try:
                    with transaction.atomic():
                        # Resolve Category
                        cat_path_str = row_data.get("Category Path")
                        cat_obj = None
                        if not is_empty(cat_path_str):
                            parts = [p.strip() for p in str(cat_path_str).split('>') if p.strip()]
                            if parts:
                                cat_obj, _ = InventoryMasterCategory.objects.get_or_create(
                                    tenant_id=tenant_id,
                                    category=parts[0],
                                    group=parts[1] if len(parts) > 1 else '',
                                    subgroup=parts[2] if len(parts) > 2 else '',
                                    sub_subgroup=parts[3] if len(parts) > 3 else ''
                                )

                        def safe_float(val, default=None):
                            if is_empty(val):
                                return default
                            try:
                                return float(str(val).replace(',', '').replace('%', '').strip())
                            except ValueError:
                                return default

                        rate_val = safe_float(row_data.get("Rate"), 0.00)
                        conv_val = safe_float(row_data.get("Conversion Factor"))
                        reorder_val = safe_float(row_data.get("Reorder Level"))
                        gst_val = safe_float(row_data.get("GST Rate"))
                        cess_val = safe_float(row_data.get("Cess Rate"))
                        is_saleable_bool = str(row_data.get("Is Saleable") or "").strip().lower() in ['yes', 'true', '1']

                        uom_str = str(row_data.get("UOM") or "nos").strip().lower()
                        alt_uom_str = str(row_data.get("Alternate UOM") or "").strip() or None

                        inv_item, created = InventoryItem.objects.update_or_create(
                            tenant_id=tenant_id,
                            item_code=str(item_code).strip(),
                            defaults={
                                "item_name": str(item_name).strip(),
                                "description": str(row_data.get("Description") or "").strip() or None,
                                "category": cat_obj,
                                "category_path": cat_obj.full_path if cat_obj else None,
                                "uom": uom_str,
                                "alternate_uom": alt_uom_str,
                                "conversion_factor": conv_val,
                                "rate": rate_val,
                                "rate_unit": str(row_data.get("Rate Unit") or uom_str).strip() or "nos",
                                "hsn_code": str(row_data.get("HSN Code") or "").strip() or None,
                                "gst_rate": gst_val,
                                "cess_rate": cess_val,
                                "reorder_level": reorder_val,
                                "is_saleable": is_saleable_bool,
                                "is_active": True,
                            }
                        )

                        results["successful_imports"].append({
                            "id": getattr(inv_item, 'id', None),
                            "item_code": inv_item.item_code,
                            "item_name": inv_item.item_name,
                            "row_data": row_data
                        })
                        results["success"] += 1

                except Exception as row_err:
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: {str(row_err)}",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    logger.error(f"Error processing confirmed inventory item row {row_idx}: {row_err}")

            return Response({
                "message": f"Import complete. Success: {results['success']}, Failed: {results['failed']}",
                "summary": results,
                "is_preview": False
            }, status=200)

        # --- PATH B: FILE UPLOAD — preview (dry_run=true) or direct file import ---
        if not excel_file:
            return Response({"error": "No file provided"}, status=400)

        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            assert ws is not None

            # Header mapping using fuzzy matching
            excel_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
            header_index = match_headers(excel_headers, INVENTORY_ITEM_COLUMNS)

            # Validate required columns — return as summary so modal shows Step 2 with error inline
            missing_cols = [col["label"] for col in INVENTORY_ITEM_COLUMNS if col["required"] and col["label"] not in header_index]
            if missing_cols:
                col_list = ", ".join(f'"{c}"' for c in missing_cols)
                return Response({
                    "summary": {
                        "success": 0,
                        "failed": 1,
                        "errors": [{
                            "message": f"Missing required column(s): {col_list}. Please use the correct template.",
                            "row_data": {},
                            "row_index": "Header"
                        }],
                        "successful_imports": []
                    },
                    "is_preview": True
                }, status=200)

            records_to_process = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row): continue
                row_data = {}
                for lbl, idx in header_index.items():
                    row_data[lbl] = row[idx-1]
                records_to_process.append({"row_data": row_data, "row_index": row_idx})

            results = {"success": 0, "failed": 0, "errors": [], "successful_imports": []}
            seen_item_codes = set()

            for item in records_to_process:
                row_data = item.get("row_data")
                row_idx = item.get("row_index", "N/A")

                if not row_data: continue

                item_code = row_data.get("Item Code")
                item_name = row_data.get("Item Name")

                def is_empty(val):
                    return not val or str(val).strip().lower() in ['n/a', 'none', 'nan', 'null', '']

                if is_empty(item_code) or is_empty(item_name):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Item Code or Item Name is missing",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                item_code_clean = str(item_code).strip()
                if item_code_clean in seen_item_codes:
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Duplicate Item Code '{item_code_clean}' in the import list",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue
                seen_item_codes.add(item_code_clean)

                if InventoryItem.objects.filter(tenant_id=tenant_id, item_code=item_code_clean, is_active=True).exists():
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Item Code '{item_code_clean}' already exists in the database",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                try:
                    with transaction.atomic():
                        # Resolve Category
                        cat_path_str = row_data.get("Category Path")
                        cat_obj = None
                        
                        if not is_empty(cat_path_str):
                            parts = [p.strip() for p in str(cat_path_str).split('>') if p.strip()]
                            if parts:
                                category = parts[0]
                                group = parts[1] if len(parts) > 1 else ''
                                subgroup = parts[2] if len(parts) > 2 else ''
                                sub_subgroup = parts[3] if len(parts) > 3 else ''
                                
                                cat_obj, _ = InventoryMasterCategory.objects.get_or_create(
                                    tenant_id=tenant_id,
                                    category=category,
                                    group=group,
                                    subgroup=subgroup,
                                    sub_subgroup=sub_subgroup
                                )

                        # Rate cleaning
                        rate_val = row_data.get("Rate")
                        if is_empty(rate_val):
                            rate_val = 0.00
                        else:
                            try:
                                rate_val = float(str(rate_val).replace(',', '').strip())
                            except ValueError:
                                rate_val = 0.00

                        # Conversion factor cleaning
                        conv_val = row_data.get("Conversion Factor")
                        if is_empty(conv_val):
                            conv_val = None
                        else:
                            try:
                                conv_val = float(str(conv_val).replace(',', '').strip())
                            except ValueError:
                                conv_val = None

                        # Reorder level cleaning
                        reorder_val = row_data.get("Reorder Level")
                        if is_empty(reorder_val):
                            reorder_val = None
                        else:
                            try:
                                reorder_val = float(str(reorder_val).replace(',', '').strip())
                            except ValueError:
                                reorder_val = None

                        # GST Rate cleaning
                        gst_val = row_data.get("GST Rate")
                        if is_empty(gst_val):
                            gst_val = None
                        else:
                            try:
                                # Remove % symbol if present
                                gst_str = str(gst_val).replace('%', '').strip()
                                gst_val = float(gst_str)
                            except ValueError:
                                gst_val = None

                        # Cess Rate cleaning
                        cess_val = row_data.get("Cess Rate")
                        if is_empty(cess_val):
                            cess_val = None
                        else:
                            try:
                                cess_str = str(cess_val).replace('%', '').strip()
                                cess_val = float(cess_str)
                            except ValueError:
                                cess_val = None

                        # Is Saleable value mapping
                        is_saleable_str = str(row_data.get("Is Saleable") or "").strip().lower()
                        is_saleable_bool = is_saleable_str in ['yes', 'true', '1']

                        # Fetch existing or instantiate new Item
                        inv_item = InventoryItem.objects.filter(
                            tenant_id=tenant_id,
                            item_code=str(item_code).strip()
                        ).first()

                        if inv_item:
                            # Update existing
                            inv_item.item_name = str(item_name).strip()
                            inv_item.description = row_data.get("Description") or None
                            inv_item.category = cat_obj
                            inv_item.category_path = cat_obj.full_path if cat_obj else None
                            inv_item.uom = str(row_data.get("UOM") or "nos").strip().lower()
                            inv_item.alternate_uom = row_data.get("Alternate UOM") or None
                            inv_item.conversion_factor = conv_val
                            inv_item.rate = rate_val
                            inv_item.rate_unit = row_data.get("Rate Unit") or row_data.get("UOM") or "nos"
                            inv_item.hsn_code = str(row_data.get("HSN Code") or "").strip() or None
                            inv_item.gst_rate = gst_val
                            inv_item.cess_rate = cess_val
                            inv_item.reorder_level = reorder_val
                            inv_item.is_saleable = is_saleable_bool
                            inv_item.is_active = True
                            inv_item.updated_by = username
                        else:
                            # Create new
                            inv_item = InventoryItem(
                                tenant_id=tenant_id,
                                item_code=str(item_code).strip(),
                                item_name=str(item_name).strip(),
                                description=row_data.get("Description") or None,
                                category=cat_obj,
                                category_path=cat_obj.full_path if cat_obj else None,
                                uom=str(row_data.get("UOM") or "nos").strip().lower(),
                                alternate_uom=row_data.get("Alternate UOM") or None,
                                conversion_factor=conv_val,
                                rate=rate_val,
                                rate_unit=row_data.get("Rate Unit") or row_data.get("UOM") or "nos",
                                hsn_code=str(row_data.get("HSN Code") or "").strip() or None,
                                gst_rate=gst_val,
                                cess_rate=cess_val,
                                reorder_level=reorder_val,
                                is_saleable=is_saleable_bool,
                                is_active=True,
                            )

                        if not dry_run:
                            inv_item.save()
                        
                        results["successful_imports"].append({
                            "id": None if dry_run else getattr(inv_item, 'id', None),
                            "item_code": inv_item.item_code,
                            "item_name": inv_item.item_name,
                            "row_data": row_data
                        })
                        results["success"] += 1

                        if dry_run:
                            raise Exception("Dry Run Rollback")
                        
                except Exception as row_err:
                    if str(row_err) == "Dry Run Rollback":
                        continue
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: {str(row_err)}",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    logger.error(f"Error processing inventory item row {row_idx}: {row_err}")

            return Response({
                "message": "Preview complete" if dry_run else f"Processing complete. Success: {results['success']}, Failed: {results['failed']}",
                "summary": results,
                "is_preview": dry_run
            }, status=200)

        except Exception as e:
            logger.error(f"Inventory Item Excel upload failed: {e}", exc_info=True)
            return Response({"error": f"Internal error: {str(e)}"}, status=500)

class InventoryItemExcelExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        tenant_id = get_tenant_from_request(request)
        wb = Workbook()
        
        ws = wb.active
        assert ws is not None
        ws.title = "Exported Items"
        
        # Headers
        for col_idx, col_def in enumerate(INVENTORY_ITEM_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["label"])
            cell.font = Font(bold=True, color="FFFFFF")
            
            bg_color = "4F46E5" if col_def.get("required") else "6B7280"
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.column_dimensions[cell.column_letter].width = 20

        # Query all items for the tenant
        items = InventoryItem.objects.filter(tenant_id=tenant_id, is_active=True).order_by('item_code')
        
        thin_side = Side(border_style="thin", color="D1D5DB")
        border_all = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        
        for row_idx, item in enumerate(items, 2):
            row_data = [
                item.item_code,
                item.item_name,
                item.description or "",
                item.category_path or (item.category.full_path if item.category else ""),
                item.uom or "nos",
                item.alternate_uom or "",
                item.conversion_factor if item.conversion_factor is not None else "",
                item.rate if item.rate is not None else 0.00,
                item.rate_unit or item.uom or "nos",
                item.hsn_code or "",
                item.gst_rate if item.gst_rate is not None else "",
                item.cess_rate if item.cess_rate is not None else "",
                item.reorder_level if item.reorder_level is not None else "",
                "Yes" if item.is_saleable else "No"
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border_all
                cell.font = Font(size=10, color="1F2937")
                
                if col_idx in [7, 8, 11, 12, 13]: # numeric columns
                    if val != "":
                        try:
                            cell.value = float(val)
                        except ValueError:
                            pass
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="inventory_items_export.xlsx"'
        wb.save(response)
        return response
