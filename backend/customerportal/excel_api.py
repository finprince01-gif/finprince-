import logging
import datetime
import re
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import io

from .database import (
    CustomerMasterCustomerBasicDetails,
    CustomerMasterCustomerGSTDetails,
    CustomerMasterCustomerBanking,
    CustomerMasterCustomerTDS,
    CustomerMasterCustomerTermsCondition,
    CustomerMasterCategory,
    CustomerMasterCustomerProductService
)
from accounting.models import MasterLedger

from core.utils import match_headers

logger = logging.getLogger(__name__)

# Column Definitions for Customer Excel
CUSTOMER_COLUMNS = [
    {"label": "Customer Name", "key": "customer_name", "required": True},
    {"label": "Customer Code", "key": "customer_code", "required": False},
    {"label": "Is Also Vendor", "key": "is_also_vendor", "required": False},
    {"label": "Category", "key": "category", "required": False},
    {"label": "PAN Number", "key": "pan_number", "required": False},
    {"label": "Contact Person", "key": "contact_person", "required": False},
    {"label": "Email Address", "key": "email_address", "required": False},
    {"label": "Contact Number", "key": "contact_number", "required": False},
    {"label": "Billing Currency", "key": "billing_currency", "required": False},
    {"label": "GST TDS Applicable", "key": "gst_tds_applicable", "required": False},
    {"label": "Registered", "key": "registered", "required": True},
    {"label": "GSTIN", "key": "gstin", "required": False},
    {"label": "Branch Name", "key": "branch_name", "required": False},
    {"label": "Address Line 1", "key": "address_line_1", "required": False},
    {"label": "Address Line 2", "key": "address_line_2", "required": False},
    {"label": "Address Line 3", "key": "address_line_3", "required": False},
    {"label": "City", "key": "city", "required": False},
    {"label": "State", "key": "state", "required": False},
    {"label": "Pincode", "key": "pincode", "required": False},
    {"label": "Country", "key": "country", "required": False},
    {"label": "Branch Contact Person", "key": "branch_contact_person", "required": False},
    {"label": "Branch Email Address", "key": "branch_email", "required": False},
    {"label": "Branch Contact Number", "key": "branch_contact_number", "required": False},
    {"label": "MSME No", "key": "msme_no", "required": False},
    {"label": "FSSAI No", "key": "fssai_no", "required": False},
    {"label": "IEC Code", "key": "iec_code", "required": False},
    {"label": "TDS Section", "key": "tds_section", "required": False},
    {"label": "TCS Enabled", "key": "tcs_enabled", "required": False},
    {"label": "TCS Section", "key": "tcs_section", "required": False},
    {"label": "Credit Period", "key": "credit_period", "required": False},
    {"label": "Credit Terms", "key": "credit_terms", "required": False},
    {"label": "Penalty Terms", "key": "penalty_terms", "required": False},
    {"label": "Delivery Terms", "key": "delivery_terms", "required": False},
    {"label": "Warranty Details", "key": "warranty_details", "required": False},
    {"label": "Force Majeure", "key": "force_majeure", "required": False},
    {"label": "Dispute Terms", "key": "dispute_terms", "required": False},
    {"label": "Bank Account No", "key": "bank_account_no", "required": False},
    {"label": "Bank Name", "key": "bank_name", "required": False},
    {"label": "IFSC Code", "key": "ifsc_code", "required": False},
    {"label": "Bank Branch", "key": "bank_branch", "required": False},
    {"label": "Swift Code", "key": "swift_code", "required": False},
    {"label": "Associated Branch", "key": "associated_branch", "required": False},
    {"label": "Item Code", "key": "item_code", "required": False},
    {"label": "Item Name", "key": "item_name", "required": False},
    {"label": "HSN/SAC Code", "key": "hsn_code", "required": False},
    {"label": "UOM", "key": "uom", "required": False},
    {"label": "Customer Item Code", "key": "cust_item_code", "required": False},
    {"label": "Customer Item Name", "key": "cust_item_name", "required": False},
    {"label": "Packing Notes", "key": "packing_notes", "required": False},
]

class CustomerExcelTemplateDownloadView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        wb = Workbook()
        
        # 1. Create Readme & Instructions sheet as the active/first sheet
        ws_readme = wb.active
        assert ws_readme is not None
        ws_readme.title = "Instructions & Readme"
        
        # Title Block
        ws_readme.merge_cells("A1:D1")
        title_cell = ws_readme["A1"]
        title_cell.value = "Customer Master Creation - Excel Upload Instructions"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_readme.row_dimensions[1].height = 40
        
        # Subtitle / Note
        ws_readme.merge_cells("A2:D2")
        sub_cell = ws_readme["A2"]
        sub_cell.value = "Please read these guidelines carefully before populating the 'Customer Template' tab to ensure a successful upload."
        sub_cell.font = Font(size=10, italic=True, color="4B5563")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_readme.row_dimensions[2].height = 20
        
        # Table Headers
        headers = ["Field Label", "Mandatory / Optional", "Format & Length", "Description & Guidelines"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_readme.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(border_style="thin", color="D1D5DB")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws_readme.row_dimensions[4].height = 25
        
        # Instruction Rows grouped by Sections matching UI Tabs
        readme_data = [
            {"section": "1. BASIC DETAILS"},
            ("Customer Name", "Mandatory", "Text", "Full legal entity name or trade name of the customer."),
            ("Category", "Mandatory", "Text", "Must match an existing customer category (e.g., Regular, VIP)."),
            ("Email Address", "Mandatory", "Email Format", "Primary communication email for sending sales invoices and payment links."),
            ("Contact Number", "Mandatory", "Numeric / Text", "Primary contact phone or mobile number."),
            ("Customer Code", "Optional", "Text", "Custom unique customer identifier. Auto-generated if left empty."),
            ("PAN Number", "Optional", "10 Characters", "Must be exactly 10 alphanumeric characters in standard PAN format (e.g., ABCDE1234F)."),
            ("Billing Currency", "Optional", "Text (e.g., INR, USD)", "Default billing currency for sales invoices. Defaults to INR if empty."),
            ("Is Also Vendor", "Optional", "Yes / No", "Indicate if this customer also acts as a supplier to your business."),
            
            {"section": "2. GST & BRANCH DETAILS"},
            ("Branch Name", "Mandatory", "Text", "Name of the customer's billing/shipping branch or location (e.g., Main Branch, HQ)."),
            ("Address Line 1", "Mandatory", "Text", "Flat/House no., building name, or street address."),
            ("Address Line 2", "Mandatory", "Text", "Locality, sector, area, or road name."),
            ("Registered", "Mandatory", "Yes / No", "Specify Yes if customer is GST registered, No if unregistered."),
            ("GSTIN", "Conditional", "15 Characters", "Mandatory if Registered is Yes. Must be exactly 15 alphanumeric characters in valid GSTIN format starting with state code."),
            ("Address Line 3, City, State, Pincode, Country", "Optional", "Text / Numeric", "Additional detailed address and geographical identifiers."),
            
            {"section": "3. PRODUCTS/SERVICES"},
            ("Item Code, Item Name", "Optional", "Text", "Initial product or service mapping requested by this customer."),
            ("HSN/SAC Code", "Optional", "Numeric", "HSN/SAC classification code. Must be entirely numeric if provided."),
            ("Customer Item Code / Name", "Optional", "Text", "Customer's internal item code and description mapping."),
            
            {"section": "4. TDS & OTHER STATUTORY DETAILS"},
            ("TDS Section", "Optional", "Text", "Applicable Income Tax TDS section for direct tax mapping."),
            ("TCS Enabled / Section", "Optional", "Text", "Specify if statutory TCS collection applies to sales invoices for this customer."),
            ("MSME / FSSAI / IEC Numbers", "Optional", "Text / Numeric", "Statutory registration numbers for compliant invoicing."),
            
            {"section": "5. BANKING INFO"},
            ("Bank Account No", "Optional", "Numeric / Text", "Customer's bank account number for digital collections/refunds."),
            ("Bank Name, Bank Branch", "Optional", "Text", "Name of the bank and specific branch location."),
            ("IFSC Code, Swift Code", "Optional", "Text / Numeric", "IFSC routing code (11 characters) or international Swift code."),
            
            {"section": "6. TERMS & CONDITIONS"},
            ("Credit Period, Credit Terms", "Optional", "Numeric / Text", "Agreed sales credit repayment window and general commercial terms."),
            ("Penalty terms, Delivery terms", "Optional", "Text", "Late payment interest clauses and shipping/freight agreements."),
            ("Warranty details, Force Majeure", "Optional", "Text", "Standard product warranty terms and liability limitation conditions."),
            ("Dispute Terms", "Optional", "Text", "Jurisdiction and arbitration rules for resolving invoice disputes.")
        ]
        
        current_row = 5
        thin_side = Side(border_style="thin", color="E5E7EB")
        border_all = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        
        for item in readme_data:
            if isinstance(item, dict) and "section" in item:
                ws_readme.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                cell = ws_readme.cell(row=current_row, column=1, value=item["section"])
                cell.font = Font(size=11, bold=True, color="1E40AF")
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                
                for c in range(1, 5):
                    ws_readme.cell(row=current_row, column=c).border = border_all
                ws_readme.row_dimensions[current_row].height = 25
                current_row += 1
            else:
                fill_color = "FFFFFF" if current_row % 2 != 0 else "F9FAFB"
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                for col_idx, val in enumerate(item, 1):
                    cell = ws_readme.cell(row=current_row, column=col_idx, value=val)
                    cell.fill = row_fill
                    cell.border = border_all
                    cell.font = Font(size=10, color="1F2937")
                    
                    if col_idx == 1:
                        cell.font = Font(size=10, bold=True, color="111827")
                        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    elif col_idx == 2:
                        is_mand = (val == "Mandatory")
                        cell.font = Font(size=10, bold=is_mand, color="B91C1C" if is_mand else "4B5563")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 3:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                ws_readme.row_dimensions[current_row].height = 22
                current_row += 1
            
        # Set specific column widths for Readme sheet
        ws_readme.column_dimensions['A'].width = 28
        ws_readme.column_dimensions['B'].width = 22
        ws_readme.column_dimensions['C'].width = 25
        ws_readme.column_dimensions['D'].width = 65
        
        # 2. Create the actual Customer Template sheet
        ws = wb.create_sheet(title="Customer Template")
        assert ws is not None
        
        # Headers
        for col_idx, col_def in enumerate(CUSTOMER_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["label"])
            cell.font = Font(bold=True, color="FFFFFF")
            
            bg_color = "3B82F6" if col_def.get("required") else "6B7280"
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.column_dimensions[cell.column_letter].width = 20

        ws.freeze_panes = "A2"
        
        # Enable sheet protection to prevent header modification
        ws.protection.sheet = True
        
        registered_col_letter = None
        gstin_col_letter = None

        # Unlock rows 2 to 1000 for data entry (up to 1000 records)
        for col_idx, col_def in enumerate(CUSTOMER_COLUMNS, 1):
            if col_def["key"] == "registered":
                registered_col_letter = get_column_letter(col_idx)
            if col_def["key"] == "gstin":
                gstin_col_letter = get_column_letter(col_idx)
                
            for row in range(2, 1001):
                ws.cell(row=row, column=col_idx).protection = Protection(locked=False)

        if registered_col_letter:
            # Dropdown for Registered
            dv_registered = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            dv_registered.error = 'Your entry is not in the list'
            dv_registered.errorTitle = 'Invalid Entry'
            dv_registered.prompt = 'Please select from the list'
            dv_registered.promptTitle = 'Select Yes/No'
            dv_registered.add(f'{registered_col_letter}2:{registered_col_letter}1000')
            ws.add_data_validation(dv_registered)

        if registered_col_letter and gstin_col_letter:
            # Custom Data Validation for GSTIN based on Registered column
            dv_gstin = DataValidation(type="custom", formula1=f'${registered_col_letter}2="Yes"', allow_blank=True)
            dv_gstin.error = 'GSTIN cannot be entered when Registered is No'
            dv_gstin.errorTitle = 'Invalid Entry'
            dv_gstin.add(f'{gstin_col_letter}2:{gstin_col_letter}1000')
            ws.add_data_validation(dv_gstin)
            
            # Conditional formatting to gray out GSTIN if Registered is "No"
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            rule = FormulaRule(formula=[f'${registered_col_letter}2="No"'], stopIfTrue=True, fill=gray_fill)
            ws.conditional_formatting.add(f'{gstin_col_letter}2:{gstin_col_letter}1000', rule)
        
        wb.active = 0
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="customer_import_template.xlsx"'
        wb.save(response)
        return response

class CustomerExcelExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        tenant_id = getattr(request.user, "tenant_id", None)
        customers = CustomerMasterCustomerBasicDetails.objects.filter(tenant_id=tenant_id, is_deleted=False)
        
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Customers"
        
        # Headers
        for col_idx, col_def in enumerate(CUSTOMER_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["label"])
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = 20

        # Data
        for row_idx, customer in enumerate(customers, 2):
            gst = customer.gst_details.first()
            banking = customer.banking_details.first()
            tds = getattr(customer, 'tds_details', None)
            terms = getattr(customer, 'terms_conditions', None)
            
            data = {
                "Customer Name": customer.customer_name,
                "Customer Code": customer.customer_code,
                "Is Also Vendor": "Yes" if customer.is_also_vendor else "No",
                "Category": str(customer.customer_category) if customer.customer_category else "",
                "PAN Number": customer.pan_number,
                "Contact Person": customer.contact_person,
                "Email Address": customer.email_address,
                "Contact Number": customer.contact_number,
                "Billing Currency": customer.billing_currency,
                "Registered": "No" if (gst and gst.is_unregistered) else "Yes",
                "GSTIN": gst.gstin if gst else "",
                "GST TDS Applicable": "Yes" if customer.gst_tds_applicable else "No",
                "Branch Name": gst.branch_reference_name if gst else "",
                "Address Line 1": gst.address_line_1 if gst else "",
                "Address Line 2": gst.address_line_2 if gst else "",
                "Address Line 3": gst.address_line_3 if gst else "",
                "City": gst.city if gst else "",
                "State": gst.state if gst else "",
                "Pincode": gst.pincode if gst else "",
                "Country": gst.country if (gst and gst.country) else "India",
                "Branch Contact Person": gst.branch_contact_person if gst else "",
                "Branch Email Address": gst.branch_email if gst else "",
                "Branch Contact Number": gst.branch_contact_number if gst else "",
                "MSME No": tds.msme_no if tds else "",
                "FSSAI No": tds.fssai_no if tds else "",
                "IEC Code": tds.iec_code if tds else "",
                "TDS Section": tds.tds_section if tds else "",
                "TCS Enabled": "Yes" if (tds and tds.tcs_enabled) else "No",
                "TCS Section": tds.tcs_section if tds else "",
                "Credit Period": terms.credit_period if terms else "",
                "Credit Terms": terms.credit_terms if terms else "",
                "Penalty Terms": terms.penalty_terms if terms else "",
                "Delivery Terms": terms.delivery_terms if terms else "",
                "Warranty Details": terms.warranty_details if terms else "",
                "Force Majeure": terms.force_majeure if terms else "",
                "Dispute Terms": terms.dispute_terms if terms else "",
                "Bank Account No": banking.account_number if banking else "",
                "Bank Name": banking.bank_name if banking else "",
                "IFSC Code": banking.ifsc_code if banking else "",
                "Bank Branch": banking.branch_name if banking else "",
            }
            
            # Fetch products
            products = CustomerMasterCustomerProductService.objects.filter(customer_basic_detail=customer).first()
            if products:
                data.update({
                    "Item Code": products.item_code,
                    "Item Name": products.item_name,
                    "HSN/SAC Code": products.hsn_code,
                    "UOM": products.uom,
                    "Customer Item Code": products.customer_item_code,
                    "Customer Item Name": products.customer_item_name,
                    "Packing Notes": products.packing_notes,
                })
            
            for col_idx, col_def in enumerate(CUSTOMER_COLUMNS, 1):
                ws.cell(row=row_idx, column=col_idx, value=data.get(str(col_def["label"]), ""))

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="customers_export.xlsx"'
        wb.save(response)
        return response

class CustomerExcelUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        tenant_id = getattr(request.user, "tenant_id", None)
        username = getattr(request.user, "username", "system")
        excel_file = request.FILES.get("file")
        json_data = request.data.get("data") # Support for sending fixed records as JSON
        dry_run = request.query_params.get("dry_run") == "true"
        
        if not excel_file and not json_data:
            return Response({"error": "No file or data provided"}, status=400)
        
        try:
            records_to_process = []
            
            if excel_file:
                wb = load_workbook(excel_file, data_only=True)
                ws = wb.active
                assert ws is not None
                
                # Header mapping using fuzzy matching
                excel_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
                header_index = match_headers(excel_headers, CUSTOMER_COLUMNS)
                
                # Validate required columns
                for col in CUSTOMER_COLUMNS:
                    if col["required"] and col["label"] not in header_index:
                        return Response({"error": f"Missing required column: {col['label']}"}, status=400)

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if not any(row): continue
                    row_data = {}
                    for lbl, idx in header_index.items():
                        row_data[lbl] = row[idx-1]
                    records_to_process.append({"row_data": row_data, "row_index": row_idx})
            else:
                # Use data from request body (for fixed records)
                if isinstance(json_data, str):
                    import json
                    records_to_process = json.loads(json_data)
                else:
                    records_to_process = json_data

            results = {"success": 0, "failed": 0, "errors": [], "successful_imports": []}
            
            # Group hierarchical rows (subsequent rows with empty customer basic details)
            grouped_records = []
            current_main_record = None

            def _raw_empty(v):
                return not v or str(v).strip().lower() in ['', 'none', 'nan', 'null', 'n/a']

            for item in records_to_process:
                r_data = item.get("row_data")
                if not r_data: continue

                c_code = r_data.get("Customer Code") or r_data.get("customer_code")
                c_name = r_data.get("Customer Name") or r_data.get("customer_name") or r_data.get("name") or r_data.get("Name")
                email = r_data.get("Email Address") or r_data.get("email_address") or r_data.get("email_id") or r_data.get("email") or r_data.get("Email ID") or r_data.get("Email")

                is_main = not (_raw_empty(c_code) and _raw_empty(c_name) and _raw_empty(email))

                if is_main:
                    if "extra_branches" not in r_data: r_data["extra_branches"] = []
                    if "extra_banks" not in r_data: r_data["extra_banks"] = []
                    if "products" not in r_data: r_data["products"] = []
                    
                    has_product = not _raw_empty(r_data.get("Item Code") or r_data.get("item_code") or r_data.get("Item Name") or r_data.get("item_name"))
                    if has_product:
                        r_data["products"].append({
                            "Item Code": r_data.get("Item Code"),
                            "Item Name": r_data.get("Item Name"),
                            "HSN/SAC Code": r_data.get("HSN/SAC Code") or r_data.get("HSN/SAC") or r_data.get("hsn_code") or r_data.get("hsn_sac_code"),
                            "UOM": r_data.get("UOM"),
                            "Customer Item Code": r_data.get("Customer Item Code") or r_data.get("Cust Item Code") or r_data.get("cust_item_code"),
                            "Customer Item Name": r_data.get("Customer Item Name") or r_data.get("Cust Item Name") or r_data.get("cust_item_name"),
                            "Packing Notes": r_data.get("Packing Notes"),
                        })
                        
                    current_main_record = item
                    grouped_records.append(item)
                else:
                    if current_main_record:
                        # Append to current main record
                        has_branch = not _raw_empty(r_data.get("Branch Name") or r_data.get("branch_name") or r_data.get("City"))
                        has_bank = not _raw_empty(r_data.get("Bank Account No") or r_data.get("bank_account_no") or r_data.get("Bank Name") or r_data.get("bank_name"))
                        has_product = not _raw_empty(r_data.get("Item Code") or r_data.get("item_code") or r_data.get("Item Name") or r_data.get("item_name"))

                        if has_branch: current_main_record["row_data"]["extra_branches"].append(r_data)
                        if has_bank: current_main_record["row_data"]["extra_banks"].append(r_data)
                        if has_product: current_main_record["row_data"]["products"].append(r_data)
                    else:
                        grouped_records.append(item)
            
            records_to_process = grouped_records

            # Cache for customers created in this file (so multiple distinct named branches in same file link together)
            file_customers_cache = {}
            file_customers_code_cache = {}

            for item in records_to_process:
                row_data = item.get("row_data")
                row_idx = item.get("row_index", "N/A")
                
                if not row_data: continue
                
                def is_empty(val):
                    if val is None: return True
                    s = str(val).strip().lower()
                    return s in ['', 'none', 'n/a', 'nan', 'null', 'n / a']

                c_name = row_data.get("Customer Name") or row_data.get("customer_name") or row_data.get("name") or row_data.get("Name")
                if is_empty(c_name):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Customer Name is missing",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue
                
                cat_name = row_data.get("Category") or row_data.get("category")
                if is_empty(cat_name):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Category is mandatory",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                email = row_data.get("Email Address") or row_data.get("email_address") or row_data.get("email_id") or row_data.get("email") or row_data.get("Email ID") or row_data.get("Email")
                if is_empty(email):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Email Address is mandatory",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                contact = row_data.get("Contact Number") or row_data.get("contact_number") or row_data.get("contact") or row_data.get("Contact No") or row_data.get("contact_no") or row_data.get("Phone") or row_data.get("phone")
                if is_empty(contact):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Contact Number is mandatory",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                branch = row_data.get("Branch Name") or row_data.get("branch_name")
                if is_empty(branch):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Branch Name is mandatory",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                addr1 = row_data.get("Address Line 1") or row_data.get("address_line_1")
                if is_empty(addr1):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Address Line 1 is mandatory",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                addr2 = row_data.get("Address Line 2") or row_data.get("address_line_2")
                if is_empty(addr2):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Address Line 2 is mandatory",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue
                
                # Check main record for Country, State, City
                country = row_data.get("Country") or row_data.get("country")
                if is_empty(country):
                    results["failed"] += 1
                    results["errors"].append({"message": f"Row {row_idx}: Country is mandatory", "row_data": row_data, "row_index": row_idx})
                    continue
                    
                state = row_data.get("State") or row_data.get("state")
                if is_empty(state):
                    results["failed"] += 1
                    results["errors"].append({"message": f"Row {row_idx}: State is mandatory", "row_data": row_data, "row_index": row_idx})
                    continue
                    
                city = row_data.get("City") or row_data.get("city")
                if is_empty(city):
                    results["failed"] += 1
                    results["errors"].append({"message": f"Row {row_idx}: City is mandatory", "row_data": row_data, "row_index": row_idx})
                    continue

                # Check extra branches
                branch_invalid = False
                for b_idx, branch_data in enumerate(row_data.get("extra_branches", [])):
                    for field_name in ["Branch Name", "Address Line 1", "Address Line 2", "Country", "State", "City"]:
                        val = branch_data.get(field_name) or branch_data.get(field_name.lower().replace(' ', '_'))
                        if is_empty(val):
                            results["failed"] += 1
                            results["errors"].append({
                                "message": f"Row {row_idx}: Branch {b_idx + 1} {field_name} is mandatory",
                                "row_data": row_data,
                                "row_index": row_idx
                            })
                            branch_invalid = True
                            break
                    if branch_invalid: break
                
                if branch_invalid: continue
                
                # Validation
                pan = row_data.get("PAN Number") or row_data.get("pan_number")
                if pan:
                    if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]{1}$', str(pan).strip().upper()):
                        results["failed"] += 1
                        results["errors"].append({
                            "message": f"Row {row_idx}: Invalid PAN format",
                            "row_data": row_data,
                            "row_index": row_idx
                        })
                        continue
                
                registered = row_data.get("Registered") or row_data.get("registered")
                if is_empty(registered):
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: Registered is mandatory (Yes/No)",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    continue

                is_registered = str(registered).strip().lower() not in ['no', 'false', '0']
                
                gstin = row_data.get("GSTIN") or row_data.get("gstin")
                if is_registered:
                    if is_empty(gstin):
                        results["failed"] += 1
                        results["errors"].append({
                            "message": f"Row {row_idx}: GSTIN is mandatory when Registered is Yes",
                            "row_data": row_data,
                            "row_index": row_idx
                        })
                        continue
                    if not re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$', str(gstin).strip().upper()):
                        results["failed"] += 1
                        results["errors"].append({
                            "message": f"Row {row_idx}: Invalid GSTIN format",
                            "row_data": row_data,
                            "row_index": row_idx
                        })
                        continue
                else:
                    if not is_empty(gstin):
                        row_data["GSTIN"] = ""
                        gstin = ""
                
                try:
                    # Use a nested transaction that we can rollback if dry_run
                    with transaction.atomic():
                        # 1. Basic Details
                        import random, string
                        cust_code = row_data.get("Customer Code")
                        
                        customer = None
                        c_name_lower = str(c_name).strip().lower()
                        
                        # Check cache first for same-file duplicates
                        if cust_code and cust_code in file_customers_code_cache:
                            raise Exception(f"DUPLICATE ENTRY: Customer Code '{cust_code}' appears more than once in this file. Each customer must have a unique code.")
                        else:
                            # Check DB for duplicate / soft-deleted customer
                            # Search by customer code
                            if cust_code:
                                existing_by_code = CustomerMasterCustomerBasicDetails.objects.filter(tenant_id=tenant_id, customer_code=cust_code).first()
                                if existing_by_code:
                                    if not existing_by_code.is_deleted:
                                        raise Exception(f"DUPLICATE ENTRY: Customer Code '{cust_code}' already exists in the system. Please use a unique code or leave it blank.")
                                    else:
                                        # It's soft-deleted, we will reactivate it
                                        customer = existing_by_code

                            if customer:
                                # Restore/reactivate soft-deleted customer
                                customer.is_deleted = False
                                customer.is_active = True
                                customer.customer_name = row_data.get("Customer Name") or row_data.get("customer_name") or row_data.get("name") or row_data.get("Name")
                                customer.customer_code = cust_code or customer.customer_code
                                customer.pan_number = row_data.get("PAN Number")
                                customer.contact_person = row_data.get("Contact Person")
                                customer.email_address = row_data.get("Email Address") or row_data.get("email_address") or row_data.get("email_id") or row_data.get("email") or row_data.get("Email ID") or row_data.get("Email")
                                customer.contact_number = row_data.get("Contact Number") or row_data.get("contact_number") or row_data.get("contact") or row_data.get("Contact No") or row_data.get("contact_no") or row_data.get("Phone") or row_data.get("phone")
                                customer.billing_currency = row_data.get("Billing Currency", "INR")
                                customer.is_also_vendor = True if str(row_data.get("Is Also Vendor", "")).lower() in ['yes', 'true', '1'] else False
                                customer.gst_tds_applicable = True if str(row_data.get("GST TDS Applicable", "")).lower() in ['yes', 'true', '1'] else False
                                
                                cat_name_val = row_data.get("Category", "Regular")
                                cat, _ = CustomerMasterCategory.objects.get_or_create(
                                    tenant_id=tenant_id,
                                    category=cat_name_val,
                                    defaults={'is_active': True}
                                )
                                customer.customer_category = cat
                                customer.updated_by = username
                                customer.save()
                                
                                # Clean up existing related records to prevent duplicate constraints or orphan records
                                CustomerMasterCustomerGSTDetails.objects.filter(customer_basic_detail=customer).delete()
                                CustomerMasterCustomerBanking.objects.filter(customer_basic_detail=customer).delete()
                                CustomerMasterCustomerTDS.objects.filter(customer_basic_detail=customer).delete()
                                CustomerMasterCustomerTermsCondition.objects.filter(customer_basic_detail=customer).delete()
                                
                                file_customers_cache[c_name_lower] = customer
                                file_customers_code_cache[customer.customer_code] = customer
                        
                        is_new_customer = False
                        if not customer:
                            is_new_customer = True
                            if not cust_code:
                                cust_code = f"CUST-{''.join(random.choices(string.ascii_uppercase + string.digits, k=6))}"
                            
                            cat_name_val = row_data.get("Category", "Regular")
                            cat, _ = CustomerMasterCategory.objects.get_or_create(
                                tenant_id=tenant_id,
                                category=cat_name_val,
                                defaults={'is_active': True}
                            )
                            
                            customer = CustomerMasterCustomerBasicDetails.objects.create(
                                tenant_id=tenant_id,
                                customer_name=row_data.get("Customer Name") or row_data.get("customer_name") or row_data.get("name") or row_data.get("Name"),
                                customer_code=cust_code,
                                pan_number=row_data.get("PAN Number"),
                                contact_person=row_data.get("Contact Person"),
                                email_address=row_data.get("Email Address") or row_data.get("email_address") or row_data.get("email_id") or row_data.get("email") or row_data.get("Email ID") or row_data.get("Email"),
                                contact_number=row_data.get("Contact Number") or row_data.get("contact_number") or row_data.get("contact") or row_data.get("Contact No") or row_data.get("contact_no") or row_data.get("Phone") or row_data.get("phone"),
                                billing_currency=row_data.get("Billing Currency", "INR"),
                                is_also_vendor=True if str(row_data.get("Is Also Vendor", "")).lower() in ['yes', 'true', '1'] else False,
                                gst_tds_applicable=True if str(row_data.get("GST TDS Applicable", "")).lower() in ['yes', 'true', '1'] else False,
                                customer_category=cat,
                                created_by=username
                            )
                            
                            # Ledger creation
                            ledger_code = f"CUST-LED-{getattr(customer, 'id', random.randint(1000, 9999))}"
                            ledger = MasterLedger.objects.create(  # type: ignore
                                tenant_id=tenant_id,
                                name=c_name,
                                group='Sundry Debtors',
                                code=ledger_code,
                            )
                            customer.ledger_id = ledger.id
                            customer.save(update_fields=['ledger_id'])
                            
                            file_customers_cache[c_name_lower] = customer
                            file_customers_code_cache[customer.customer_code] = customer
                        
                        # 2. GST Details
                        all_branches = [row_data] + row_data.get("extra_branches", [])
                        for branch_data in all_branches:
                            gstin = branch_data.get("GSTIN")
                            registered_str = str(branch_data.get("Registered") or branch_data.get("registered") or "").strip().lower()
                            is_unregistered = registered_str in ['no', 'false', '0']
                            branch_ref = branch_data.get("Branch Name", "Main Branch")

                            if is_new_customer or not CustomerMasterCustomerGSTDetails.objects.filter(
                                tenant_id=tenant_id, customer_basic_detail=customer, branch_reference_name__iexact=branch_ref
                            ).exists():
                                CustomerMasterCustomerGSTDetails.objects.create(  # type: ignore
                                    tenant_id=tenant_id,
                                    customer_basic_detail=customer,
                                    gstin=gstin if not is_unregistered else None,
                                    is_unregistered=is_unregistered,
                                    branch_reference_name=branch_ref,
                                    address_line_1=branch_data.get("Address Line 1"),
                                    address_line_2=branch_data.get("Address Line 2"),
                                    address_line_3=branch_data.get("Address Line 3"),
                                    city=branch_data.get("City"),
                                    state=branch_data.get("State"),
                                    pincode=branch_data.get("Pincode"),
                                    country=branch_data.get("Country") or branch_data.get("country") or "India",
                                    branch_contact_person=branch_data.get("Branch Contact Person") or branch_data.get("branch_contact_person"),
                                    branch_email=branch_data.get("Branch Email Address") or branch_data.get("branch_email"),
                                    branch_contact_number=branch_data.get("Branch Contact Number") or branch_data.get("branch_contact_number"),
                                    created_by=username
                                )
                            
                        # 3. Banking Details
                        all_banks = [row_data] + row_data.get("extra_banks", [])
                        for bank_data in all_banks:
                            acc_no = bank_data.get("Bank Account No")
                            if acc_no:
                                if is_new_customer or not CustomerMasterCustomerBanking.objects.filter(
                                    tenant_id=tenant_id, customer_basic_detail=customer, account_number=acc_no
                                ).exists():
                                    CustomerMasterCustomerBanking.objects.create(  # type: ignore
                                        tenant_id=tenant_id,
                                        customer_basic_detail=customer,
                                        account_number=acc_no,
                                        bank_name=bank_data.get("Bank Name"),
                                        ifsc_code=bank_data.get("IFSC Code"),
                                        branch_name=bank_data.get("Bank Branch"),
                                        swift_code=bank_data.get("Swift Code"),
                                        associated_branches=[bank_data.get("Associated Branch")] if bank_data.get("Associated Branch") else [],
                                        created_by=username
                                    )
                            elif bank_data is row_data and is_new_customer:
                                CustomerMasterCustomerBanking.objects.create(tenant_id=tenant_id, customer_basic_detail=customer, created_by=username)  # type: ignore
 
                        # 4. TDS Details (Only create if new customer or doesn't exist)
                        if is_new_customer or not CustomerMasterCustomerTDS.objects.filter(tenant_id=tenant_id, customer_basic_detail=customer).exists():
                            CustomerMasterCustomerTDS.objects.create(  # type: ignore
                                tenant_id=tenant_id,
                                customer_basic_detail=customer,
                                msme_no=row_data.get("MSME No"),
                                fssai_no=row_data.get("FSSAI No"),
                                iec_code=row_data.get("IEC Code"),
                                tds_section=row_data.get("TDS Section"),
                                tcs_section=row_data.get("TCS Section"),
                                tcs_enabled=True if str(row_data.get("TCS Enabled", "")).lower() in ['yes', 'true', '1'] else False,
                                created_by=username
                            )
                        
                        # 5. Terms & conditions (Only create if new customer or doesn't exist)
                        if is_new_customer or not CustomerMasterCustomerTermsCondition.objects.filter(tenant_id=tenant_id, customer_basic_detail=customer).exists():
                            CustomerMasterCustomerTermsCondition.objects.create(  # type: ignore
                                tenant_id=tenant_id,
                                customer_basic_detail=customer,
                                credit_period=row_data.get("Credit Period"),
                                credit_terms=row_data.get("Credit Terms"),
                                penalty_terms=row_data.get("Penalty Terms"),
                                delivery_terms=row_data.get("Delivery Terms"),
                                warranty_details=row_data.get("Warranty Details"),
                                force_majeure=row_data.get("Force Majeure"),
                                dispute_terms=row_data.get("Dispute Terms"),
                                created_by=username
                            )
                        
                        # 6. Products/Services
                        products = row_data.get("products", [])
                        if not products:
                            # Try to get from flat fields
                            item_code = row_data.get("Item Code") or row_data.get("item_code")
                            if item_code:
                                products = [{
                                    "Item Code": item_code,
                                    "Item Name": row_data.get("Item Name") or row_data.get("item_name"),
                                    "HSN/SAC Code": row_data.get("HSN/SAC Code") or row_data.get("HSN/SAC") or row_data.get("hsn_code") or row_data.get("hsn_sac_code"),
                                    "UOM": row_data.get("UOM") or row_data.get("uom"),
                                    "Customer Item Code": row_data.get("Customer Item Code") or row_data.get("Cust Item Code") or row_data.get("cust_item_code"),
                                    "Customer Item Name": row_data.get("Customer Item Name") or row_data.get("Cust Item Name") or row_data.get("cust_item_name"),
                                    "Packing Notes": row_data.get("Packing Notes") or row_data.get("packing_notes"),
                                }]
                        
                        for p in products:
                            item_code = p.get("Item Code") or p.get("item_code")
                            if not item_code: continue
                            
                            hsn = p.get("HSN/SAC Code") or p.get("HSN/SAC") or p.get("hsn_code") or p.get("hsn_sac_code")
                            if hsn and not str(hsn).replace(" ", "").isdigit():
                                raise Exception(f"Invalid HSN/SAC Code: '{hsn}'. HSN codes must be numeric.")

                            if is_new_customer or not CustomerMasterCustomerProductService.objects.filter(
                                tenant_id=tenant_id, customer_basic_detail=customer, item_code=item_code
                            ).exists():
                                CustomerMasterCustomerProductService.objects.create(  # type: ignore
                                    tenant_id=tenant_id,
                                    customer_basic_detail=customer,
                                    item_code=item_code,
                                    item_name=p.get("Item Name") or p.get("item_name"),
                                    hsn_code=hsn,
                                    uom=p.get("UOM") or p.get("uom"),
                                    customer_item_code=p.get("Customer Item Code") or p.get("Cust Item Code") or p.get("cust_item_code") or p.get("customer_item_code"),
                                    customer_item_name=p.get("Customer Item Name") or p.get("Cust Item Name") or p.get("cust_item_name") or p.get("customer_item_name"),
                                    packing_notes=p.get("Packing Notes") or p.get("packing_notes"),
                                    created_by=username
                                )
                        
                        results["successful_imports"].append({
                            "id": None if dry_run else getattr(customer, 'id', None),
                            "name": customer.customer_name,
                            "code": customer.customer_code,
                            "row_data": row_data # Include row data for preview editing
                        })
                        results["success"] += 1

                        if dry_run:
                            # Rollback this specific row
                            raise Exception("Dry Run Rollback")
                        
                except Exception as row_err:
                    if str(row_err) == "Dry Run Rollback":
                        continue # This is expected for dry run, we already counted success
                    results["failed"] += 1
                    results["errors"].append({
                        "message": f"Row {row_idx}: {str(row_err)}",
                        "row_data": row_data,
                        "row_index": row_idx
                    })
                    logger.error(f"Error processing customer row {row_idx}: {row_err}")

            return Response({
                "message": "Preview complete" if dry_run else f"Processing complete. Success: {results['success']}, Failed: {results['failed']}",
                "summary": results,
                "is_preview": dry_run
            }, status=200)

        except Exception as e:
            logger.error(f"Customer Excel upload failed: {e}", exc_info=True)
            return Response({"error": f"Internal error: {str(e)}"}, status=500)
