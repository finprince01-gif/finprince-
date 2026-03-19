import os
import io
import json
import tempfile
import uuid
import base64
import traceback

from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.core.cache import cache
from django.db import transaction
from vendors.models import VendorMasterBasicDetail, VendorMasterGSTDetails
from accounting.serializers_voucher_purchase import VoucherPurchaseSupplierDetailsSerializer
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# Configure Gemini
GEN_API_KEY = os.environ.get('GEMINI_API_KEY')
if GEN_API_KEY:
    genai.configure(api_key=GEN_API_KEY)

# Header-level fields (appear once per invoice, not per line item)
HEADER_FIELDS = [
    # Voucher Details
    'Voucher Date', 'Supplier Invoice No', 'Purchase Voucher No', 'Vendor Name', 'GSTIN', 'PAN', 'MSME Number',
    'Bill From', 'Ship From', 'Input Type', 'Invoice in Foreign Currency', 'Supporting Document',
    
    # Dispatch Details
    'Dispatch From', 'Mode of Transport', 'Dispatch Date', 'Dispatch Time', 'Delivery Type',
    'Self/Third Party', 'Transporter ID', 'Transporter Name', 'Vehicle No', 'LR/GR Consignment', 'Dispatch Document',
    
    # e-Way Bill Details
    'e-Way Bill Available', 'e-Way Bill No', 'e-Way Bill Date', 'Validity Period', 'Distance',
    'Extension Date', 'Extended EWB No', 'Extension Reason', 'From Place', 'Remaining Distance',
    'New Validity', 'Updated Vehicle No', 'IRN', 'Ack No',
    
    # Total Details
    'Total Taxable Value', 'Total IGST', 'Total CGST', 'Total SGST', 'Total Cess', 'Total State Cess',
    'Total Invoice Value', 'TDS Income Tax', 'TDS GST', 'Advance Paid', 'To Pay', 'Posting Note', 'Terms & Conditions',
    'Advance References',

    # Additional Required Fields
    'GST Registration',
    'Original Sales Invoice Value for Credit Note',
    'e-Invoice - Ack No.',
    'e-Invoice - Ack Date',
    'e-Invoice - IRN',
    'e-Invoice - Bill to place',
    'e-Invoice - Ship to place',
    'e-Invoice - Dispatch From Name',
    'e-Invoice - Dispatch From Address',
    'e-Invoice - Dispatch From State',
    'e-Invoice - Dispatch From Pincode',
    'e-Invoice - Dispatch From Place',
    'e-Invoice Cancellation - Reason for Cancellation',
    'e-Invoice Cancellation - Remarks',
    'e-Way Bill No.',
    'e-Way Bill Date',
    'Consolidated e-Way Bill No.',
    'Consolidated e-Way Bill Date',
    'e-Way Bill - Sub-Type',
    'e-Way Bill - Document Type',
    'Consignor Details (From) - Address-1',
    'Consignor Details (From) - Address-2',
    'Consignor Details (From) - Address Type',
    'Consignor Details (From) - Pincode',
    'Consignor Details (From) - Place',
    'Consignor Details (From) - Actual State',
    'Consignee Details (To) - Address-1',
    'Consignee Details (To) - Address-2',
    'Consignee Details (To) - Address Type',
    'Consignee Details (To) - Place',
    'Consignee Details (To) - Actual State',
    'Consignee Details (To) - Pincode',
    'e-Way Bill Transport Details - Pin to Pin Distance as per Portal',
    'e-Way Bill Transport Details - Transporter Name',
    'e-Way Bill Transport Details - Transporter ID',
    'e-Way Bill Transport Details - Mode',
    'e-Way Bill Transport Details - Doc/Lading/RR/Airway No.',
    'e-Way Bill Transport Details - Doc/Lading/RR/Airway Date',
    'e-Way Bill Transport Details - Vehicle Number',
    'e-Way Bill Transport Details - Vehicle Type',
    'e-Way Bill Transport Details - Place',
    'e-Way Bill Transport Details - State',
    'e-Way Bill Transport Details - Reason',
    'e-Way Bill Transport Details - Remarks',
    'e-Way Bill Extension Details - Remaining Distance(in KM)',
    'e-Way Bill Extension Details - Mode',
    'e-Way Bill Extension Details - Doc/Lading/RR/Airway No.',
    'e-Way Bill Extension Details - Doc/Lading/RR/Airway - Date',
    'e-Way Bill Extension Details - Vehicle Number',
    'e-Way Bill Extension Details - Vehicle Type',
    'e-Way Bill Extension Details - Transit Type',
    'e-Way Bill Extension Details - Address 1',
    'e-Way Bill Extension Details - Address 2',
    'e-Way Bill Extension Details - Address 3',
    'e-Way Bill Extension Details - Current Pincode',
    'e-Way Bill Extension Details - Current Place',
    'e-Way Bill Extension Details - Current State',
    'e-Way Bill Extension Details - Reason',
    'e-Way Bill Extension Details - Remarks',
    'e-Way Bill Cancellation Details - Reason',
    'e-Way Bill Cancellation Details - Remarks',
    'GST Rate Details',
    'GST Source of Details',
    'GST Source Type of Master',
    'GST Taxability Type',
    'GST Nature of Transaction',
    'GST Classification',
    'IGST Rate',
    'CGST Rate',
    'SGST/UTGST Rate',
    'Cess Rate',
    'Cess Rate Per Unit',
    'State Cess Rate',
    'Applicable for Reverse Charge',
    'Eligible for Input Tax Credit',
    'Taxable Value',
    'HSN/SAC Details',
    'HSN/SAC Source of Details',
    'HSN/SAC Source Type of Master',
    'HSN/SAC Classification',
    'HSN/SAC',
    'HSN Description',
    'Buyer/Supplier - Bill to/from',
    'Buyer/Supplier - Address Type',
    'Buyer/Supplier - Mailing Name',
    'Buyer/Supplier - Address',
    'Buyer/Supplier - Country',
    'Buyer/Supplier - State',
    'Buyer/Supplier - GST Registration Type',
    'Buyer/Supplier - Assessee of Other Territory',
    'Buyer/Supplier - GSTIN/UIN',
    'Buyer/Supplier - Is Bill of Entry available',
    'Buyer/Supplier - Supplies under section 7 of IGST Act',
    'Buyer/Supplier - Place of Supply',
    'Consignee (ship to)',
    'Consignee - Mailing Name',
    'Consignee - Address Type',
    'Consignee - Address',
    'Consignee - State',
    'Consignee - Country',
    'Consignee - GSTIN/UIN',
    'Stat Adjustment (GST) - Type of Duty/Tax',
    'Stat Adjustment (GST) - Nature of Adjustment',
    'Stat Adjustment (GST) - Additional Nature of Adjustment',
    'Stat Adjustment (GST) - Rate',
    'Stat Adjustment (GST) - Taxable Value',
    'Stat Adjustment (GST) - ISD Invoice/Debit/Credit Note No.',
    'Stat Adjustment (GST) - ISD Invoice/Debit/Credit Note Date',
    'Stat Adjustment (GST) - Eligible for Input Tax Credit',
    'Stat Adjustment (GST) - Type of Supply',
    'Type of Supply',
    'Advance Payment/Receipt/Refund Details - IGST Rate',
    'Advance Payment/Receipt/Refund Details - IGST Amount',
    'Advance Payment/Receipt/Refund Details - SGST Rate',
    'Advance Payment/Receipt/Refund Details - SGST Amount',
    'Advance Payment/Receipt/Refund Details - CGST Rate',
    'Advance Payment/Receipt/Refund Details - CGST Amount',
    'Advance Payment/Receipt/Refund Details - Cess Rate',
    'Advance Payment/Receipt/Refund Details - Cess Amount',
    'Advance Payment/Receipt/Refund Details - Cess Rate Per Unit',
    'Advance Payment/Receipt/Refund Details - Cess per Unit Amount',
    'Tax Type Allocations - IGST Liability',
    'Tax Type Allocations - CGST Liability',
    'Tax Type Allocations - SGST/UTGST Liability',
    'Tax Type Allocations - Cess Liability',
    'GST Advance Details - Month Year',
    'GST Advance Details - Place of Supply',
    'GST Advance Details - GST Rate',
    'GST Advance Details - Cess Rate',
    'GST Advance Details - Advance Amount',
    'TDS - Nature of Payments',
    'TDS - Assessable Value',
    'TDS Party Details - Party Name',
    'TDS Party Details - Deductee Type',
    'TDS Party Details - PAN Number',
    'TDS Bill Allocations - Type of Ref',
    'TDS Bill Allocations - Name',
    'TDS Bill Allocations - TDS Nature of Payment',
    'TDS Bill Allocations - Party Ledger',
    'TDS Bill Allocations - Expenses Ledger',
    'TDS Bill Allocations - Duty Ledger',
    'TDS Bill Allocations - Assessable Amount',
    'TDS Bill Allocations - Payable Amount',
    'TDS Bill Allocations - Paid Amount',
    'TCS - Nature of Goods',
    'TCS - Assessable Value',
    'Exemption from TCS for Buyer-Deductible TDS',
    'TCS Party Details - Party Name',
    'TCS Party Details - Collectee Type',
    'TCS Party Details - PAN Number',
    'TCS Bill Allocations - Type of Ref',
    'TCS Bill Allocations - Name',
    'TCS Bill Allocations - TCS Nature of Goods',
    'TCS Bill Allocations - Party Ledger',
    'TCS Bill Allocations - Income Ledger',
    'TCS Bill Allocations - Duty Ledger',
    'TCS Bill Allocations - Assessable Amount',
    'TCS Bill Allocations - Payable Amount',
    'TCS Bill Allocations - Paid Amount',
    'Stat Payment (GST) - Tax Type',
    'Stat Payment (GST) - Type of Payment',
    'Stat Payment (GST) - Period From',
    'Stat Payment (GST) - Period To',
    'Stat Payment (TDS) - Tax Type',
    'Stat Payment (TDS) - Period From',
    'Stat Payment (TDS) - Period To',
    'Stat Payment (TDS) - Section',
    'Stat Payment (TDS) - Nature of Payment',
    'Stat Payment (TDS) - Deductee Status',
    'Stat Payment (TDS) - Residential Status',
    'Stat Payment (TDS) - Cheque No.',
    'Stat Payment (TDS) - Cheque Date',
    'Stat Payment (TDS) - BSR Code',
    'Stat Payment (TDS) - Challan No.',
    'Stat Payment (TDS) - Challan Date',
    'Stat Payment (TDS) - Bank Name',
    'Stat Payment (TDS) - Branch Name',
    'Stat Payment (TCS) - Tax Type',
    'Stat Payment (TCS) - Period From',
    'Stat Payment (TCS) - Period To',
    'Stat Payment (TCS) - Section',
    'Stat Payment (TCS) - Nature of Goods',
    'Stat Payment (TCS) - Deductee Status',
    'Stat Payment (TCS) - Residential Status',
    'Stat Payment (TCS) - Cheque No.',
    'Stat Payment (TCS) - Cheque Date',
    'Stat Payment (TCS) - BSR Code',
    'Stat Payment (TCS) - Challan No.',
    'Stat Payment (TCS) - Challan Date',
    'Stat Payment (TCS) - Bank Name',
    'Stat Payment (TCS) - Branch Name'
]

# Per-line-item fields (one value per printed row in the items table)
LINE_ITEM_FIELDS = [
    'S.No', 'Item Name', 'Purchase Ledger', 'HSN/SAC', 'Quantity', 'UOM', 'Rate', 
    'Disc %', 'Disc Amount', 'Taxable Value', 'GST %', 'Integrated Tax (IGST)', 
    'Central Tax (CGST)', 'State Tax (SGST)', 'Cess', 'Item Amount', 'Description'
]

# All column headers in final Excel (line-item fields first, then header fields)
ALL_HEADERS = LINE_ITEM_FIELDS + HEADER_FIELDS


@csrf_exempt
@require_http_methods(["POST"])
def extract_invoice(request):
    """
    1. Accept uploaded invoice image / PDF.
    2. If PDF is multi-invoice, split it into per-invoice temp PDFs first.
    3. Run Gemini OCR + mapping engine on each detected invoice.
    4. Return an array of results (one per invoice).
       For a single-page PDF or image, returns a single-element result.
    """
    if not request.FILES.get('file'):
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    uploaded_file = request.FILES['file']

    try:
        from core.ai_proxy import execute_with_retry, api_key_manager

        # Get healthy key
        api_key = api_key_manager.get_healthy_key()
        if not api_key:
            return JsonResponse({'error': 'AI service busy (No healthy keys)'}, status=503)

        # Read file bytes once
        file_bytes = uploaded_file.read()
        mime_type = uploaded_file.content_type or 'application/octet-stream'
        original_name = uploaded_file.name or 'invoice'

        # ── Multi-invoice PDF splitting (preprocessing) ──────────────────────
        is_pdf = (
            mime_type == 'application/pdf'
            or (original_name or '').lower().endswith('.pdf')
        )

        if is_pdf:
            from core.pdf_splitter import split_pdf_into_invoice_files, cleanup_temp_pdf
            invoice_chunks = split_pdf_into_invoice_files(
                pdf_bytes=file_bytes,
                original_filename=original_name,
            )
        else:
            # Non-PDF (image): treat as a single chunk
            invoice_chunks = [('SINGLE', None, None)]  # sentinel

        # ── Process each detected invoice through the existing OCR pipeline ──
        import concurrent.futures

        def process_chunk(chunk):
            inv_number, tmp_path, group = chunk
            try:
                # Load the bytes for this invoice
                if tmp_path:                   # PDF split chunk
                    with open(tmp_path, 'rb') as fh:
                        chunk_bytes = fh.read()
                    chunk_mime = 'application/pdf'
                    chunk_label = (
                        f"{original_name} [Inv {inv_number}]"
                        if len(invoice_chunks) > 1
                        else original_name
                    )
                else:                          # Non-PDF image
                    chunk_bytes = file_bytes
                    chunk_mime = mime_type
                    chunk_label = original_name

                # ── Existing prompt (unchanged) ──────────────────────────────
                prompt_text = f"""
You are a precision invoice OCR and data-extraction system.
Extract every figure exactly where it is printed — correct column, correct row — with zero shifting, duplication, or guessing.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PHASE 1 — LOCK COLUMN BOUNDARIES (do this first)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Find the table header row (S.No, Description, HSN, Qty, Rate, Amount, etc.).
2. Record the left/right boundary of each column. LOCK them.
3. Typical order: S.No | Item Name | Purchase Ledger | HSN/SAC | Qty | UOM | Rate | Disc% | Taxable Amt | GST% | IGST | CGST | SGST | Cess | Amount

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SEMANTIC MAPPING GUIDE (Strict Schema Enforcement)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Every extracted field MUST be assigned to its matching header key.

- Voucher Date                <- The invoice date printed on the document.
- Supplier Invoice No         <- The main invoice/bill number.
- Vendor Name                 <- The seller/company providing the invoice.
- GSTIN                       <- The seller's GST number.
- PAN                         <- The seller's PAN (if printed).
- Bill From                   <- The seller's full address.
- Ship From                   <- The shipping origin address.
- Total Invoice Value         <- The final grand total payable amount.
- Total Taxable Value         <- The sum of pre-tax item amounts.
- Total IGST/CGST/SGST        <- Corresponding invoice-level tax totals.

- Item Name                   <- The main product name or description. MANDATORY.
- Description                 <- Extra details ONLY.
- Quantity                    <- The numeric unit count.
- UOM                         <- The unit abbreviation (PCS, BOX, KGS, etc.).
- Rate                        <- The unit price.
- Taxable Value (Item)        <- Item pre-tax total (Qty x Rate).
- GST %                       <- The tax percentage for that item row.

For any field not explicitly printed, leave as "".
Never shift a value into the wrong column key.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PHASE 2 — EXTRACT BY LOCKED COLUMN (for every data row)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
For every row, read the value that sits inside each locked column boundary.
Rules:
  • A value belongs to a column ONLY if its center falls INSIDE that zone.
  • No dynamic key creation. Use provided keys only.
  • Ambiguous? -> null.
  • Strict header-based mapping — NO positional shifting.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FIELD RULES (mandatory for every row)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Item Name
  → The main product name or description. MANDATORY.
  → Map "Description", "Particulars", or "Item Name" to this field.

HSN/SAC
  → Map the 4-8 digit code here.
Quantity     → numeric part only. "8 NOS" → "8".
UOM          → unit part only. "8 NOS" → "NOS". Never combine with Quantity.
Rate         → per-unit price from Rate column only. Not the line total.
Taxable Value → pre-tax line total for THIS row only (Qty × Rate − disc). Not invoice total.
Item Amount  → final line total from Amount column for THIS row only.

ABSOLUTE PROHIBITIONS:
✗ NEVER put Grand Total / Invoice Total into any line_item field.
✗ NEVER swap Rate and Amount columns.
✗ NEVER merge two printed rows into one object.
✗ NEVER concatenate descriptions from different rows.
✗ NEVER duplicate a row.
✗ NEVER carry values from previous row into next row.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ROW RULES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Each visually distinct printed data row = ONE object in "line_items".
2. Description wrapping → attach ALL wrapped text to the "Description" field of the same object.
3. Summary rows (Sub Total, Grand Total, Tax Summary) → "invoice" fields ONLY, not items.
4. S.No increments 1, 2, 3 … by actual distinct item rows.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HEADER RULES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- "Bill From" & "Ship From" → join all address lines with ", ".
- Dates → dd/mm/yyyy. All values → strings. Missing → "".

Return the data in this EXACT JSON structure:

{{
  "invoice": {{
    {', '.join([f'"{f}": ""' for f in HEADER_FIELDS])}
  }},
  "items": [
    {{
      "S.No": "1",
      "Item Name": "",
      "Purchase Ledger": "",
      "HSN/SAC": "",
      "Quantity": "",
      "UOM": "",
      "Rate": "",
      "Disc %": "",
      "Disc Amount": "",
      "Taxable Value": "",
      "GST %": "",
      "Integrated Tax (IGST)": "",
      "Central Tax (CGST)": "",
      "State Tax (SGST)": "",
      "Cess": "",
      "Item Amount": "",
      "Description": ""
    }}
  ]
}}
Return ONLY the raw JSON object. No markdown, no code fences, no explanation.
"""

                raw_text = execute_with_retry(
                    [
                        prompt_text,
                        {'mime_type': chunk_mime, 'data': chunk_bytes}
                    ],
                    {},
                    api_key
                )

                # ── Existing mapping engine (unchanged) ───────────────────────
                from core.processing_engine import parse_and_process_ocr
                try:
                    processed_data = parse_and_process_ocr(raw_text)
                    invoice_data = processed_data.get('invoice', {})
                    items = processed_data.get('items', [])
                except Exception as e:
                    return {
                        'error': f'Failed to parse AI response: {str(e)}',
                        'source_file': chunk_label,
                    }

                # Filter empty header fields for UI display
                filtered_header = {k: v for k, v in invoice_data.items() if v and str(v).strip()}

                # ── Build Excel per invoice ───────────────────────────────────
                _tmp_dir = tempfile.gettempdir()
                _file_name = (
                    f"Invoice_Export_{uuid.uuid4().hex[:8]}_"
                    f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                _file_path = os.path.join(_tmp_dir, _file_name)

                wb = Workbook()
                ws = wb.active
                ws.title = "Extracted Invoice"

                header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)

                for col, header in enumerate(ALL_HEADERS, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)

                _items = items if items else [{}]
                for row_idx, item in enumerate(_items, start=2):
                    for col, field in enumerate(ALL_HEADERS, start=1):
                        if field in HEADER_FIELDS:
                            value = invoice_data.get(field, '')
                        else:
                            value = item.get(field, '')
                        ws.cell(row=row_idx, column=col, value=value)

                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

                wb.save(_file_path)
                with open(_file_path, "rb") as f:
                    excel_base64 = base64.b64encode(f.read()).decode('utf-8')
                if os.path.exists(_file_path):
                    os.remove(_file_path)

                return {
                    'success': True,
                    'source_file': chunk_label,
                    'invoice_number': inv_number if len(invoice_chunks) > 1 else None,
                    'data': {
                        'header': filtered_header,
                        'line_items': items,
                    },
                    'excel_file': excel_base64,
                    'file_name': _file_name,
                }

            except Exception as chunk_err:
                return {
                    'error': str(chunk_err),
                    'source_file': chunk_label if 'chunk_label' in dir() else original_name,
                    'trace': traceback.format_exc(),
                }
            finally:
                # Clean up the temp split PDF
                if tmp_path:
                    from core.pdf_splitter import cleanup_temp_pdf
                    cleanup_temp_pdf(tmp_path)
                
                from django.db import connection
                connection.close()

        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            all_invoice_results = list(executor.map(process_chunk, invoice_chunks))

        # ── Return ───────────────────────────────────────────────────────────
        if len(all_invoice_results) == 1:
            # Single invoice: preserve original flat response shape for backward compat
            result = all_invoice_results[0]
            if 'error' in result:
                return JsonResponse({'error': result['error']}, status=500)
            return JsonResponse({
                'success': True,
                'data': result['data'],
                'excel_file': result['excel_file'],
                'file_name': result['file_name'],
            })
        else:
            # Multiple invoices detected: return array so frontend can create one row per invoice
            return JsonResponse({
                'success': True,
                'multi_invoice': True,
                'invoice_count': len(all_invoice_results),
                'results': all_invoice_results,
                # Backward compat: expose first invoice's data at top-level
                'data': all_invoice_results[0].get('data') if all_invoice_results else {},
                'excel_file': all_invoice_results[0].get('excel_file') if all_invoice_results else '',
                'file_name': all_invoice_results[0].get('file_name') if all_invoice_results else '',
            })

    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)


# --------------------------------------------------------------------------------
# BULK SCAN PROCESS (3-STEP)
# --------------------------------------------------------------------------------

def detect_vendor(tenant_id, gstin, vendor_name):
    """
    Helper to find vendor by GSTIN or Name.
    """
    if gstin:
        gst_record = VendorMasterGSTDetails.objects.filter(
            tenant_id=tenant_id,
            gstin__iexact=gstin.strip()
        ).select_related('vendor_basic_detail').first()
        if gst_record:
            return gst_record.vendor_basic_detail, "FOUND"

    if vendor_name:
        vendor = VendorMasterBasicDetail.objects.filter(
            tenant_id=tenant_id,
            vendor_name__iexact=vendor_name.strip()
        ).first()
        if vendor:
            return vendor, "FOUND"
            
    return None, "MISSING"

@csrf_exempt
@require_http_methods(["POST"])
def bulk_scan_invoices(request):
    """
    STEP 1 - BULK SCAN (EXTRACTION ONLY)
    """
    files = request.FILES.getlist('files')
    if not files:
        return JsonResponse({'error': 'No files uploaded'}, status=400)

    tenant_id = getattr(request.user, 'tenant_id', 'default_tenant')
    scan_id = str(uuid.uuid4())
    results = []
    
    from core.ai_proxy import execute_with_retry, api_key_manager

    for uploaded_file in files:
        try:
            # Get healthy key
            api_key = api_key_manager.get_healthy_key()
            if not api_key:
                results.append({
                    'file_name': uploaded_file.name,
                    'error': 'AI service busy',
                    'vendor_status': 'FAILED'
                })
                continue

            file_bytes = uploaded_file.read()
            
            # Prompt (Simplified version of extract_invoice prompt)
            prompt_text = f"""
Extract invoice data in JSON:
{{
  "invoice": {{
    "Voucher Date": "", "Supplier Invoice No": "", "Vendor Name": "", "GSTIN": "",
    "Total Invoice Value": "", "Total Taxable Value": "", "Total IGST": "", "Total CGST": "", "Total SGST": ""
  }},
  "items": [
    {{ "Item Name": "", "Quantity": "", "UOM": "", "Rate": "", "Taxable Value": "", "GST %": "" }}
  ]
}}
"""
            raw_text = execute_with_retry([prompt_text, {'mime_type': uploaded_file.content_type, 'data': file_bytes}], {}, api_key)
            raw_text = raw_text.strip().strip('```json').strip('```').strip()
            extracted_json = json.loads(raw_text)
            
            invoice_data = extracted_json.get('invoice', {})
            items = extracted_json.get('items', [])
            
            vendor_name = invoice_data.get('Vendor Name', '')
            gstin = invoice_data.get('GSTIN', '')
            
            vendor, status = detect_vendor(tenant_id, gstin, vendor_name)
            
            results.append({
                'file_name': uploaded_file.name,
                'vendor_status': status,
                'vendor_id': vendor.id if vendor else None,
                'vendor_name': vendor.vendor_name if vendor else vendor_name,
                'gstin': vendor.gstin if (vendor and hasattr(vendor, 'gstin')) else gstin,
                'extracted_data': extracted_json
            })
            
        except Exception as e:
            results.append({
                'file_name': uploaded_file.name,
                'error': str(e),
                'vendor_status': 'FAILED'
            })

    # Store in cache for 1 hour
    cache.set(f"bulk_scan_{scan_id}", results, timeout=3600)
    
    # Response to frontend
    frontend_results = []
    for res in results:
        frontend_results.append({
            'file_name': res.get('file_name'),
            'vendor_status': res.get('vendor_status'),
            'vendor_id': res.get('vendor_id'),
            'vendor_name': res.get('vendor_name'),
            'gstin': res.get('gstin'),
            'error': res.get('error')
        })

    return JsonResponse({
        'scan_id': scan_id,
        'results': frontend_results
    })

@csrf_exempt
@require_http_methods(["POST"])
def resolve_bulk_vendor(request):
    """
    STEP 2 - UPDATE CACHE WITH NEW VENDOR ID
    """
    data = json.loads(request.body)
    scan_id = data.get('scan_id')
    file_name = data.get('file_name')
    vendor_id = data.get('vendor_id')
    
    if not scan_id or not file_name or not vendor_id:
        return JsonResponse({'error': 'Missing required fields'}, status=400)
        
    cache_key = f"bulk_scan_{scan_id}"
    results = cache.get(cache_key)
    if not results:
        return JsonResponse({'error': 'Scan session expired or invalid'}, status=404)
        
    updated = False
    for res in results:
        if res.get('file_name') == file_name:
            res['vendor_id'] = vendor_id
            res['vendor_status'] = 'RESOLVED'
            updated = True
            break
            
    if updated:
        cache.set(cache_key, results, timeout=3600)
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'error': 'File not found in scan results'}, status=404)

@csrf_exempt
@require_http_methods(["POST"])
def bulk_finalize_vouchers(request):
    """
    STEP 3 - FINALIZE & SAVE VOUCHERS
    """
    data = json.loads(request.body)
    scan_id = data.get('scan_id')
    
    if not scan_id:
        return JsonResponse({'error': 'Missing scan_id'}, status=400)
        
    cache_key = f"bulk_scan_{scan_id}"
    results = cache.get(cache_key)
    if not results:
        return JsonResponse({'error': 'Scan session expired or invalid'}, status=404)
        
    tenant_id = getattr(request.user, 'tenant_id', 'default_tenant')
    
    summary = {
        'total_processed': len(results),
        'created_count': 0,
        'failed_count': 0,
        'errors': []
    }
    
    for res in results:
        if res.get('vendor_status') == 'MISSING' or res.get('vendor_status') == 'FAILED':
            summary['failed_count'] = int(summary['failed_count']) + 1
            summary['errors'].append({
                'file_name': res.get('file_name'),
                'error': 'Vendor unresolved or scan failed'
            })
            continue
            
        try:
            extracted_data = res.get('extracted_data')
            invoice = extracted_data.get('invoice', {})
            items = extracted_data.get('items', [])
            
            def clean_num(v):
                if not v: return 0
                import re
                s = re.sub(r'[^\d.]', '', str(v))
                return float(s) if s else 0

            raw_date = invoice.get('Voucher Date', '')
            try:
                if '/' in raw_date:
                    d, m, y = raw_date.split('/')
                    formatted_date = f"{y}-{m}-{d}"
                else:
                    formatted_date = raw_date
            except:
                formatted_date = datetime.now().strftime('%Y-%m-%d')

            payload = {
                'date': formatted_date,
                'supplier_invoice_no': invoice.get('Supplier Invoice No', ''),
                'vendor_id': res.get('vendor_id'),
                'gstin': invoice.get('GSTIN', ''),
                'tenant_id': tenant_id,
                'supply_inr_details': {
                    'items': [
                        {
                            'item_name': item.get('Item Name', ''),
                            'qty': clean_num(item.get('Quantity', 0)),
                            'uom': item.get('UOM', ''),
                            'rate': clean_num(item.get('Rate', 0)),
                            'taxable_value': clean_num(item.get('Taxable Value', 0)),
                            'gst_pct': clean_num(item.get('GST %', 0)),
                        } for item in items
                    ]
                },
                'due_details': {
                    'to_pay': clean_num(invoice.get('Total Invoice Value', 0)),
                    'tds_gst': clean_num(invoice.get('Total IGST', 0)) + clean_num(invoice.get('Total CGST', 0)) + clean_num(invoice.get('Total SGST', 0))
                }
            }
            
            with transaction.atomic():
                serializer = VoucherPurchaseSupplierDetailsSerializer(data=payload, context={'request': request})
                if serializer.is_valid():
                    instance = serializer.save(tenant_id=tenant_id)
                    summary['created_count'] = int(summary['created_count']) + 1
                else:
                    summary['failed_count'] = int(summary['failed_count']) + 1
                    summary['errors'].append({
                        'file_name': res.get('file_name'),
                        'error': serializer.errors
                    })
        except Exception as e:
            summary['failed_count'] = int(summary['failed_count']) + 1
            summary['errors'].append({
                'file_name': res.get('file_name'),
                'error': str(e)
            })
            
    cache.delete(cache_key)
    return JsonResponse(summary)
