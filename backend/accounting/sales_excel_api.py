import logging
import datetime
from decimal import Decimal
from django.core.cache import cache
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
import uuid
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from .serializers_voucher_sales import VoucherSalesInvoiceDetailsSerializer

logger = logging.getLogger(__name__)

from .sales_validation_logic import validate_sales_customer_and_invoice
import io
import json
import base64
from inventory.models import InventoryItem
from services.models import Service
from core.models import Tenant

# Single Source of Truth for Sales Voucher Columns
# Mirrors frontend src/constants/salesVoucherColumns.ts exactly
SALES_VOUCHER_COLUMNS = [
    # ── Tab 1: Invoice Details ──────────────────────────────────────────────
    {"label": "Date",                      "key": "date",                  "tab": "Invoice Details",              "required": False, "type": "date"},
    {"label": "Sales Invoice Series",      "key": "voucher_name",          "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Sales Invoice No.",         "key": "sales_invoice_no",      "tab": "Invoice Details",              "required": True,  "type": "string"},
    {"label": "Outward Slip No.",          "key": "outward_slip_no",       "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Customer Name",             "key": "customer_name",         "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Branch",                    "key": "customer_branch",       "tab": "Invoice Details",              "required": True,  "type": "string"},
    {"label": "GSTIN",                     "key": "gstin",                 "tab": "Invoice Details",              "required": True,  "type": "string"},
    {"label": "Contact",                   "key": "contact",               "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Place of Supply",           "key": "place_of_supply",       "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Reverse Charge",            "key": "reverse_charge",        "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Nature of Supply",          "key": "invoice_type",          "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Exchange Rate",             "key": "exchange_rate",         "tab": "Invoice Details",              "required": False, "type": "number"},

    # Bill To
    {"label": "Bill To - Address Line 1",  "key": "bill_to_address_1",     "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Bill To - Address Line 2",  "key": "bill_to_address_2",     "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Bill To - Address Line 3",  "key": "bill_to_address_3",     "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Bill To - City",            "key": "bill_to_city",          "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Bill To - State",           "key": "bill_to_state",         "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Bill To - Pincode",         "key": "bill_to_pincode",       "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Bill To - Country",         "key": "bill_to_country",       "tab": "Invoice Details",              "required": False, "type": "string"},

    # Ship To
    {"label": "Ship To - Address Line 1",  "key": "ship_to_address_1",     "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Ship To - Address Line 2",  "key": "ship_to_address_2",     "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Ship To - Address Line 3",  "key": "ship_to_address_3",     "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Ship To - City",            "key": "ship_to_city",          "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Ship To - State",           "key": "ship_to_state",         "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Ship To - Pincode",         "key": "ship_to_pincode",       "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Ship To - Country",         "key": "ship_to_country",       "tab": "Invoice Details",              "required": False, "type": "string"},

    # Export Details
    {"label": "Export Type",               "key": "export_type",           "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Port Code",                 "key": "port_code",             "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Shipping Bill Number",      "key": "shipping_bill_number",  "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "Shipping Bill Date",        "key": "shipping_bill_date",    "tab": "Invoice Details",              "required": False, "type": "date"},
    {"label": "E-Commerce Operator",       "key": "ecommerce_operator",    "tab": "Invoice Details",              "required": False, "type": "string"},
    {"label": "E-Commerce GSTIN",          "key": "ecommerce_gstin",       "tab": "Invoice Details",              "required": False, "type": "string"},

    # ── Tab 2: Item & Tax Details ───────────────────────────────────────────
    {"label": "Sales Order/Quotation No.",   "key": "sales_order_no",        "tab": "Item & Tax Details",           "required": False, "type": "string"},
    {"label": "Item Code",                 "key": "item_code",             "tab": "Item & Tax Details",           "required": False, "type": "string"},
    {"label": "Item Name",                 "key": "item_name",             "tab": "Item & Tax Details",           "required": False, "type": "string"},
    {"label": "HSN / SAC",                 "key": "hsn_sac",               "tab": "Item & Tax Details",           "required": False, "type": "string"},
    {"label": "Quantity",                  "key": "qty",                   "tab": "Item & Tax Details",           "required": False, "type": "number"},
    {"label": "UQC / UOM",                 "key": "uom",                   "tab": "Item & Tax Details",           "required": False, "type": "string"},
    {"label": "Alternate Unit",            "key": "alternate_unit",        "tab": "Item & Tax Details",           "required": False, "type": "string"},
    {"label": "Rate",                      "key": "item_rate",             "tab": "Item & Tax Details",           "required": False, "type": "number"},
    {"label": "Taxable Value",             "key": "taxable_value",         "tab": "Item & Tax Details",           "required": False, "type": "number"},
    {"label": "CGST",                      "key": "cgst",                  "tab": "Item & Tax Details",           "required": False, "type": "number"},
    {"label": "SGST",                      "key": "sgst",                  "tab": "Item & Tax Details",           "required": False, "type": "number"},
    {"label": "IGST",                      "key": "igst",                  "tab": "Item & Tax Details",           "required": False, "type": "number"},
    {"label": "Cess",                      "key": "cess",                  "tab": "Item & Tax Details",           "required": False, "type": "number"},
    {"label": "Invoice Value",             "key": "invoice_value",         "tab": "Item & Tax Details",           "required": False, "type": "number"},
    {"label": "Sales Ledger",              "key": "sales_ledger",          "tab": "Item & Tax Details",           "required": False, "type": "string"},
    {"label": "ledger narration",          "key": "description",           "tab": "Item & Tax Details",           "required": False, "type": "string"},

    # ── Tab 2b: Foreign Currency (Item & Tax Details) ─────────────────────
    {"label": "Billing Currency",              "key": "fc_billing_currency", "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "string"},
    {"label": "Exchange Rate (FC to INR)",      "key": "fc_exchange_rate",    "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "number"},
    {"label": "FC - Item Name",                "key": "fc_item_name",        "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "string"},
    {"label": "FC - Quantity",                 "key": "fc_qty",              "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "number"},
    {"label": "FC - UQC / UOM",                "key": "fc_uom",              "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "string"},
    {"label": "FC - Alternate Unit",           "key": "fc_alternate_unit",   "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "string"},
    {"label": "FC - Rate (Foreign Currency)",  "key": "fc_item_rate",        "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "number"},
    {"label": "FC - Amount (Foreign Currency)", "key": "fc_invoice_value",    "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "number"},
    {"label": "FC - Sales Ledger",             "key": "fc_sales_ledger",     "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "string"},
    {"label": "FC - ledger narration",     "key": "fc_description",      "tab": "Foreign Currency (Item & Tax Details)", "required": False, "type": "string"},

    # ── Tab 3: Payment Details ──────────────────────────────────────────────
    {"label": "Total Taxable Value",       "key": "payment_taxable_value", "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "Total CGST",                "key": "payment_cgst",          "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "Total SGST / UTGST",        "key": "payment_sgst",          "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "Total IGST",                "key": "payment_igst",          "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "Total Cess",                "key": "payment_cess",          "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "State Cess",                "key": "payment_state_cess",    "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "Grand Total",               "key": "payment_invoice_value", "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "TDS (Income Tax)",          "key": "payment_tds_income_tax","tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "TDS (GST)",                 "key": "payment_tds_gst",       "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "Advance Adjusted",          "key": "payment_advance",       "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "Net Payable",               "key": "payment_payable",       "tab": "Payment Details",              "required": False, "type": "number"},
    {"label": "Posting Note",              "key": "posting_note",          "tab": "Payment Details",              "required": False, "type": "string"},
    {"label": "Terms & Conditions",        "key": "terms_conditions",      "tab": "Payment Details",              "required": False, "type": "string"},

    # ── Tab 4: Dispatch Details ─────────────────────────────────────────────
    {"label": "Dispatch From",             "key": "dispatch_from",         "tab": "Dispatch Details",             "required": False, "type": "string"},
    {"label": "Mode of Transport",         "key": "mode_of_transport",     "tab": "Dispatch Details",             "required": False, "type": "string"},
    {"label": "Dispatch Date",             "key": "dispatch_date",         "tab": "Dispatch Details",             "required": False, "type": "date"},
    {"label": "Dispatch Time",             "key": "dispatch_time",         "tab": "Dispatch Details",             "required": False, "type": "string"},
    {"label": "Delivery Type",             "key": "delivery_type",         "tab": "Dispatch Details",             "required": False, "type": "string"},
    {"label": "Self / Third Party",        "key": "self_third_party",      "tab": "Dispatch Details",             "required": False, "type": "string"},
    {"label": "Transporter ID",            "key": "transporter_id",        "tab": "Dispatch Details",             "required": False, "type": "string"},
    {"label": "Transporter Name",          "key": "transporter_name",      "tab": "Dispatch Details",             "required": False, "type": "string"},
    {"label": "Vehicle No",                "key": "vehicle_no",            "tab": "Dispatch Details",             "required": False, "type": "string"},
    {"label": "LR / GR / Consignment No",  "key": "lr_gr_consignment",     "tab": "Dispatch Details",             "required": False, "type": "string"},

    {"label": "Upto Port - Shipping Bill No",   "key": "upto_port_shipping_bill_no",   "tab": "Dispatch Details",   "required": False, "type": "string"},
    {"label": "Upto Port - Shipping Bill Date", "key": "upto_port_shipping_bill_date", "tab": "Dispatch Details",   "required": False, "type": "date"},
    {"label": "Upto Port - Port Code",          "key": "upto_port_ship_port_code",     "tab": "Dispatch Details",   "required": False, "type": "string"},
    {"label": "Upto Port - Origin",             "key": "upto_port_origin",             "tab": "Dispatch Details",   "required": False, "type": "string"},

    {"label": "Beyond Port - Shipping Bill No",  "key": "beyond_port_shipping_bill_no",  "tab": "Dispatch Details",  "required": False, "type": "string"},
    {"label": "Beyond Port - Shipping Bill Date","key": "beyond_port_shipping_bill_date","tab": "Dispatch Details",  "required": False, "type": "date"},
    {"label": "Beyond Port - Port Code",         "key": "beyond_port_ship_port_code",    "tab": "Dispatch Details",  "required": False, "type": "string"},
    {"label": "Beyond Port - Vessel / Flight No","key": "beyond_port_vessel_flight_no",  "tab": "Dispatch Details",  "required": False, "type": "string"},
    {"label": "Beyond Port - Port of Loading",   "key": "beyond_port_port_of_loading",   "tab": "Dispatch Details",  "required": False, "type": "string"},
    {"label": "Beyond Port - Port of Discharge", "key": "beyond_port_port_of_discharge", "tab": "Dispatch Details",  "required": False, "type": "string"},
    {"label": "Beyond Port - Final Destination", "key": "beyond_port_final_destination", "tab": "Dispatch Details",  "required": False, "type": "string"},
    {"label": "Beyond Port - Origin",            "key": "beyond_port_origin",            "tab": "Dispatch Details",  "required": False, "type": "string"},
    {"label": "Beyond Port - Origin Country",    "key": "beyond_port_origin_country",    "tab": "Dispatch Details",  "required": False, "type": "string"},
    {"label": "Beyond Port - Destination Country","key": "beyond_port_dest_country",      "tab": "Dispatch Details",  "required": False, "type": "string"},

    {"label": "Rail Upto Port - Delivery Type",     "key": "rail_upto_port_delivery_type",     "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Upto Port - Transporter ID",    "key": "rail_upto_port_transporter_id",    "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Upto Port - Transporter Name",  "key": "rail_upto_port_transporter_name",  "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Upto Port - Vehicle No",        "key": "rail_upto_port_vehicle_no",        "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Upto Port - LR / GR / Consignment No", "key": "rail_upto_port_lr_gr",    "tab": "Dispatch Details", "required": False, "type": "string"},

    {"label": "Rail Beyond Port - Railway Receipt No",   "key": "rail_beyond_port_receipt_no",   "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Beyond Port - Railway Receipt Date", "key": "rail_beyond_port_receipt_date", "tab": "Dispatch Details", "required": False, "type": "date"},
    {"label": "Rail Beyond Port - Origin",               "key": "rail_beyond_port_origin",       "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Beyond Port - Origin Country",       "key": "rail_beyond_port_origin_country","tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Beyond Port - Rail No",              "key": "rail_beyond_port_rail_no",      "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Beyond Port - FNR No",               "key": "rail_beyond_port_fnr_no",       "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Beyond Port - Station of Loading",   "key": "rail_beyond_port_station_loading", "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Beyond Port - Station of Discharge", "key": "rail_beyond_port_station_discharge", "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Beyond Port - Final Destination",   "key": "rail_beyond_port_final_destination", "tab": "Dispatch Details", "required": False, "type": "string"},
    {"label": "Rail Beyond Port - Destination Country",  "key": "rail_beyond_port_dest_country", "tab": "Dispatch Details", "required": False, "type": "string"},

    # ── Tab 5: E-Invoice & E-Way Bill Details ─────────────────────────────
    {"label": "IRN",                       "key": "irn",                   "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "Acknowledgement No",        "key": "ack_no",                "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "Acknowledgement Date",      "key": "ack_date",              "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "date"},

    # Entry 1
    {"label": "EWB 1 - Available",         "key": "ewb1_available",        "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 1 - E-Way Bill No",      "key": "ewb1_eway_bill_no",     "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 1 - Date",              "key": "ewb1_date",             "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "date"},
    {"label": "EWB 1 - Validity Period",   "key": "ewb1_validity_period",  "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 1 - Distance (km)",     "key": "ewb1_distance",         "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "number"},
    {"label": "EWB 1 - Extension Date",    "key": "ewb1_extension_date",   "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "date"},
    {"label": "EWB 1 - Extended EWB No",   "key": "ewb1_extended_ewb_no",  "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 1 - Extension Reason",  "key": "ewb1_extension_reason", "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 1 - From Place",        "key": "ewb1_from_place",       "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 1 - Remaining Distance","key": "ewb1_remaining_distance","tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "number"},
    {"label": "EWB 1 - New Validity",      "key": "ewb1_new_validity",     "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 1 - Updated Vehicle No","key": "ewb1_updated_vehicle_no","tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},

    # Entry 2
    {"label": "EWB 2 - Available",         "key": "ewb2_available",        "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 2 - E-Way Bill No",      "key": "ewb2_eway_bill_no",     "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 2 - Date",              "key": "ewb2_date",             "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "date"},
    {"label": "EWB 2 - Validity Period",   "key": "ewb2_validity_period",  "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 2 - Distance (km)",     "key": "ewb2_distance",         "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "number"},
    {"label": "EWB 2 - Extension Date",    "key": "ewb2_extension_date",   "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "date"},
    {"label": "EWB 2 - Extended EWB No",   "key": "ewb2_extended_ewb_no",  "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 2 - Extension Reason",  "key": "ewb2_extension_reason", "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 2 - From Place",        "key": "ewb2_from_place",       "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 2 - Remaining Distance","key": "ewb2_remaining_distance","tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "number"},
    {"label": "EWB 2 - New Validity",      "key": "ewb2_new_validity",     "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 2 - Updated Vehicle No","key": "ewb2_updated_vehicle_no","tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},

    # Entry 3
    {"label": "EWB 3 - Available",         "key": "ewb3_available",        "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 3 - E-Way Bill No",      "key": "ewb3_eway_bill_no",     "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 3 - Date",              "key": "ewb3_date",             "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "date"},
    {"label": "EWB 3 - Validity Period",   "key": "ewb3_validity_period",  "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 3 - Distance (km)",     "key": "ewb3_distance",         "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "number"},
    {"label": "EWB 3 - Extension Date",    "key": "ewb3_extension_date",   "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "date"},
    {"label": "EWB 3 - Extended EWB No",   "key": "ewb3_extended_ewb_no",  "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 3 - Extension Reason",  "key": "ewb3_extension_reason", "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 3 - From Place",        "key": "ewb3_from_place",       "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 3 - Remaining Distance","key": "ewb3_remaining_distance","tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "number"},
    {"label": "EWB 3 - New Validity",      "key": "ewb3_new_validity",     "tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
    {"label": "EWB 3 - Updated Vehicle No","key": "ewb3_updated_vehicle_no","tab": "E-Invoice & E-Way Bill Details", "required": False, "type": "string"},
]

HEADER_LABELS = [c["label"] for c in SALES_VOUCHER_COLUMNS]
REQUIRED_LABELS = [c["label"] for c in SALES_VOUCHER_COLUMNS if c.get("required")]

TAB_COLORS = {
    "Invoice Details": "3B82F6",                       # Blue
    "Item & Tax Details": "10B981",                    # Emerald
    "Foreign Currency (Item & Tax Details)": "F59E0B", # Amber
    "Payment Details": "6366F1",                       # Indigo
    "Dispatch Details": "8B5CF6",                      # Violet
    "E-Invoice & E-Way Bill Details": "EC4899"         # Pink
}

class SalesVoucherColumnSchemaView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        return Response({"columns": SALES_VOUCHER_COLUMNS})

class SalesExcelTemplateDownloadView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        if ws is None: ws = wb.create_sheet()
        ws.title = "Sales Voucher"
        
        # Headers
        for col_idx, col_def in enumerate(SALES_VOUCHER_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["label"])
            cell.font = Font(bold=True, color="FFFFFF")
            
            tab_name = str(col_def["tab"])
            bg_color = TAB_COLORS.get(tab_name, "000000")
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            # Width
            ws.column_dimensions[cell.column_letter].width = max(len(str(col_def["label"])) + 5, 15)

        ws.freeze_panes = "A2"
        
        # Enable sheet protection to prevent header modification
        ws.protection.sheet = True
        
        # Unlock rows 2 to 1000 for data entry (up to 1000 records)
        for row in range(2, 1001):
            for col in range(1, len(SALES_VOUCHER_COLUMNS) + 1):
                ws.cell(row=row, column=col).protection = Protection(locked=False)
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="sales_voucher_template.xlsx"'
        wb.save(response)
        return response

class SalesExcelErrorReportView(APIView):
    """
    Generates a downloadable Excel report for failed rows.
    Accepts a list of errors with row data in the POST body.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        errors = request.data.get("errors", [])
        if not errors:
            return Response({"error": "No error data provided"}, status=status.HTTP_400_BAD_REQUEST)

        wb = Workbook()
        ws = wb.active
        if ws is None: ws = wb.create_sheet()
        ws.title = "Failed Rows"

        # Headers: Original Labels + Error Message
        headers = HEADER_LABELS + ["Error Reason"]
        for idx, lbl in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=lbl)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red

        # Data
        for row_idx, err_obj in enumerate(errors, 2):
            row_dict = err_obj.get("row_data", {})
            err_msg = err_obj.get("errors", {})
            if isinstance(err_msg, dict):
                # Flatten dict errors
                msg_str = ""
                for k, v in err_msg.items():
                    msg_str += f"{k}: {v}; "
                err_msg = msg_str
            
            for col_idx, label in enumerate(HEADER_LABELS, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_dict.get(label))
            
            # Error Message in last column
            ws.cell(row=row_idx, column=len(headers), value=str(err_msg))

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="sales_upload_errors.xlsx"'
        wb.save(response)
        return response

class SalesExcelUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        tenant_id = getattr(request.user, "tenant_id", 1)
        excel_file = request.FILES.get("file")
        
        if not excel_file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            if ws is None:
                return Response({"error": "No active sheet found in Excel file"}, status=status.HTTP_400_BAD_REQUEST)
            
            # 1. Map Headers to Columns (Case-insensitive & Trimmed)
            # Create a index map using normalized Uppercase labels
            excel_col_index = {}
            for idx, cell in enumerate(ws[1]):
                if cell.value:
                    lbl = str(cell.value).strip().upper()
                    excel_col_index[lbl] = idx
            
            # Map of Our internal Label -> Excel Column Index
            header_map = {}
            for col_cfg in SALES_VOUCHER_COLUMNS:
                label = str(col_cfg["label"])
                norm = label.strip().upper()
                if norm in excel_col_index:
                    header_map[label] = excel_col_index[norm]

            uploaded_labels = {str(k).upper() for k in excel_col_index.keys()}
            missing_cols = [str(l) for l in REQUIRED_LABELS if str(l).strip().upper() not in uploaded_labels]
            if missing_cols:
                return Response({
                    "error": "Missing required columns",
                    "missing_columns": missing_cols
                }, status=status.HTTP_400_BAD_REQUEST)

            # 2. Extract Data Rows
            rows_data = []
            if ws is None: rows_data = [] # Extra safety
            else:
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if not any(row): continue # Skip empty rows
                    
                    row_dict = {}
                    for label, col_idx in header_map.items():
                        row_dict[label] = row[col_idx] if col_idx < len(row) else None
                    
                    row_dict["_row_num"] = row_idx
                    rows_data.append(row_dict)

            if not rows_data:
                return Response({"error": "No data found in Excel file"}, status=status.HTTP_400_BAD_REQUEST)

            v_no_label = "Sales Invoice No." # This matches our updated label above
            vouchers_groups = {}
            last_v_no = None
            for rd in rows_data:
                v_no = str(rd.get(v_no_label) or "").strip()
                if v_no:
                    last_v_no = v_no
                
                if not last_v_no: continue
                
                if last_v_no not in vouchers_groups:
                    vouchers_groups[last_v_no] = []
                
                if not v_no:
                    rd["Sales Invoice No."] = last_v_no # Inherit
                
                vouchers_groups[last_v_no].append(rd)

            # 4. Processing Helpers
            def _safe_float(v):
                if str(v).startswith("="): v = str(v).replace("=", "")
                if str(v).startswith("+"): v = str(v).replace("+", "")
                if str(v).startswith("-"): v = str(v).replace("-", "")
                if v is None or v == "": return 0.0
                try: 
                    if isinstance(v, str):
                        v = v.replace(",", "").strip()
                    return float(v)
                except: return 0.0

            def _get_date(v, default=None):
                if v is None or v == "": return default
                if hasattr(v, 'date'): # Handles datetime.datetime and potentially other objects
                    return v.date()
                if isinstance(v, datetime.date):
                    return v
                if isinstance(v, str) and v.strip():
                    try:
                        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
                            try: return datetime.datetime.strptime(v.strip(), fmt).date()
                            except: continue
                    except: pass
                if isinstance(v, (int, float)):
                    try:
                        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=v))
                    except: return default or datetime.date.today()
                return default or datetime.date.today()

            def _get_time(v):
                if v is None or v == "": return None
                if isinstance(v, datetime.time):
                    return v.strftime("%H:%M:%S")
                if isinstance(v, datetime.datetime):
                    return v.time().strftime("%H:%M:%S")
                if isinstance(v, (int, float)):
                    try:
                        total_seconds = round(v * 24 * 3600)
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        return datetime.time(hours % 24, minutes, seconds).strftime("%H:%M:%S")
                    except: return None
                if isinstance(v, str) and v.strip():
                    v_str = v.strip()
                    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
                        try:
                            clean_v = v_str.replace(".", ":") if "." in v_str and ":" not in v_str else v_str
                            return datetime.datetime.strptime(clean_v, fmt).time().strftime("%H:%M:%S")
                        except: continue
                return None

            # 5. Process Each Group
            created_vouchers = []
            creation_errors = []
            total_items_count = 0 # Initialize counter
            # Pre-fetch inventory and service masters for auto-filling item details
            inventory_master = list(InventoryItem.objects.filter(tenant_id=tenant_id, is_active=True).values( # type: ignore
                'item_code', 'item_name', 'hsn_code', 'uom', 'rate', 'gst_rate'
            ))
            service_master = list(Service.objects.filter(tenant_id=tenant_id, is_active=True).values( # type: ignore
                'service_code', 'service_name', 'sac_code', 'uom', 'gst_rate'
            ))

            for v_no, group_rows in vouchers_groups.items():
                first = group_rows[0]
                
                # Extract Header (Invoice Details)
                invoice_payload = {}
                for col in SALES_VOUCHER_COLUMNS:
                    if col["tab"] == "Invoice Details":
                        val = first.get(col["label"])
                        key = col["key"]
                        
                        if col["type"] == "date":
                            invoice_payload[key] = _get_date(val)
                        elif col["type"] == "number":
                            invoice_payload[key] = _safe_float(val)
                        else:
                            # If blank, set to None (NULL in DB)
                            val_str = str(val).strip() if (val is not None and str(val).strip() != "") else None
                            
                            # Normalization for short fields
                            if key == "reverse_charge" and val_str:
                                first_char = val_str[0].upper()
                                val_str = "Y" if first_char == "Y" else "N"

                            # Restore critical defaults to avoid DRF/DB validation errors for non-nullable fields
                            if val_str is None:
                                if key == "reverse_charge": val_str = "N"
                                elif key == "invoice_type": val_str = "Regular"
                                elif key == "exchange_rate": val_str = "1"
                                elif key == "customer_name": val_str = "Imported Customer"

                            invoice_payload[key] = val_str

                # Compulsory Validation Check as per requirement
                missing_mandatory = []
                if not first.get("Sales Invoice No."): missing_mandatory.append("Sales Invoice No.")
                if not first.get("Branch"):            missing_mandatory.append("Branch")
                if not first.get("GSTIN"):             missing_mandatory.append("GSTIN")
                if not first.get("Customer Name"):     missing_mandatory.append("Customer Name")

                if missing_mandatory:
                    creation_errors.append({
                        "row_num": first.get("_row_num"),
                        "voucher_number": v_no or "Unknown",
                        "errors": {"validation_error": [f"Compulsory fields missing: {', '.join(missing_mandatory)}"]},
                        "status": "Failed"
                    })
                    continue
                
                # Address Mapping
                s_order_no = str(first.get("Sales Order No.") or "").strip()
                if s_order_no:
                    invoice_payload["sales_order_no"] = s_order_no

                bill_to_parts = [first.get(f"Bill To - Address Line {i}") for i in range(1, 4)] + \
                                [first.get("Bill To - City"), first.get("Bill To - State"), 
                                 first.get("Bill To - Pincode"), first.get("Bill To - Country")]
                invoice_payload["bill_to"] = "\n".join(filter(None, [str(p).strip() for p in bill_to_parts if p]))
                
                ship_to_parts = [first.get(f"Ship To - Address Line {i}") for i in range(1, 4)] + \
                                [first.get("Ship To - City"), first.get("Ship To - State"), 
                                 first.get("Ship To - Pincode"), first.get("Ship To - Country")]
                invoice_payload["ship_to"] = "\n".join(filter(None, [str(p).strip() for p in ship_to_parts if p]))

                # 4. Extract Items from ALL rows in group
                items = []
                foreign_items = []
                for row_dict in group_rows:
                    item_obj = {}
                    has_data = False
                    for col in SALES_VOUCHER_COLUMNS:
                        if col["tab"] == "Item & Tax Details":
                            if col["key"] == "sales_order_no": continue
                            val = row_dict.get(col["label"])
                            if val not in (None, ""):
                                has_data = True
                                item_obj[col["key"]] = _safe_float(val) if col["type"] == "number" else str(val).strip()
                            else:
                                item_obj[col["key"]] = 0.0 if col["type"] == "number" else None
                    if has_data:
                        # Auto-fill details from Master if missing
                        code = str(item_obj.get("item_code") or "").strip()
                        name = str(item_obj.get("item_name") or "").strip()
                        
                        master_item = None
                        if code:
                            master_item = next((m for m in inventory_master if str(m['item_code']).strip().upper() == code.upper()), None)
                            if not master_item:
                                master_item = next((m for m in service_master if str(m['service_code']).strip().upper() == code.upper()), None)
                        
                        if not master_item and name:
                            master_item = next((m for m in inventory_master if str(m['item_name']).strip().upper() == name.upper()), None)
                            if not master_item:
                                master_item = next((m for m in service_master if str(m['service_name']).strip().upper() == name.upper()), None)
                        
                        if master_item:
                            # Fill missing basic info
                            if not item_obj.get("item_code"): item_obj["item_code"] = master_item.get("item_code") or master_item.get("service_code")
                            if not item_obj.get("item_name") or item_obj["item_name"] == "Imported Item": 
                                item_obj["item_name"] = master_item.get("item_name") or master_item.get("service_name")
                            if not item_obj.get("hsn_sac"): item_obj["hsn_sac"] = master_item.get("hsn_code") or master_item.get("sac_code")
                            if not item_obj.get("uom"): item_obj["uom"] = master_item.get("uom")
                            
                            # Fill Rate if missing or zero
                            m_rate = _safe_float(master_item.get("rate") or 0)
                            if (not item_obj.get("item_rate") or item_obj["item_rate"] == 0) and m_rate > 0:
                                item_obj["item_rate"] = m_rate
                                
                            # Calculate Taxable Value if missing
                            qty = _safe_float(item_obj.get("qty") or 0)
                            rate = _safe_float(item_obj.get("item_rate") or 0)
                            if (not item_obj.get("taxable_value") or item_obj["taxable_value"] == 0) and qty > 0 and rate > 0:
                                item_obj["taxable_value"] = round(qty * rate, 2)
                            
                            # Auto-calculate Taxes if master has GST Rate and columns are empty
                            m_gst = _safe_float(master_item.get("gst_rate") or 0)
                            taxable = _safe_float(item_obj.get("taxable_value") or 0)
                            
                            has_any_tax = any(_safe_float(item_obj.get(k) or 0) > 0 for k in ["cgst", "sgst", "igst", "cess"])
                            
                            if m_gst > 0 and taxable > 0 and not has_any_tax:
                                # Determine if Intra-state or Inter-state
                                pos = str(invoice_payload.get("place_of_supply") or "").strip().lower()
                                
                                # Fetch company state from tenant
                                tenant_obj = Tenant.objects.filter(id=tenant_id).first() # type: ignore
                                company_state = str(tenant_obj.state or "tamil nadu").strip().lower()
                                
                                total_tax = round(taxable * (m_gst / 100), 2)
                                
                                if pos and company_state and company_state not in pos:
                                    item_obj["igst"] = total_tax
                                else:
                                    item_obj["cgst"] = round(total_tax / 2, 2)
                                    item_obj["sgst"] = round(total_tax / 2, 2)

                        if not item_obj.get("item_name"): item_obj["item_name"] = "Imported Item"
                        items.append(item_obj)
                    
                    fc_obj = {}
                    has_fc_data = False
                    fc_mapping = {
                        "FC - Item Name": "item_name", "FC - Quantity": "quantity", "FC - UQC / UOM": "uqc",
                        "FC - Alternate Unit": "alternate_unit", "FC - Rate (Foreign Currency)": "rate",
                        "FC - Amount (Foreign Currency)": "amount", "FC - Sales Ledger": "sales_ledger", "FC - Item Narration": "description"
                    }
                    for label, key in fc_mapping.items():
                        val = row_dict.get(label)
                        if val not in (None, ""):
                            has_fc_data = True
                            fc_obj[key] = _safe_float(val) if ("Rate" in label or "Amount" in label or "Quantity" in label) else str(val).strip()
                        else:
                            fc_obj[key] = None if ("Rate" not in label and "Amount" not in label and "Quantity" not in label) else 0.0
                    if has_fc_data:
                        if not fc_obj.get("item_name"): fc_obj["item_name"] = "Imported FC Item"
                        foreign_items.append(fc_obj)
                
                print(f"DEBUG BRAIN: Voucher {v_no} -> extracted {len(items)} items and {len(foreign_items)} FC items")
                invoice_payload["items"] = items
                invoice_payload["foreign_items"] = foreign_items

                # 5. Extraction: Payment, Dispatch, EWB
                payment_details = {}
                for col in SALES_VOUCHER_COLUMNS:
                    if col["tab"] == "Payment Details":
                        val = first.get(col["label"])
                        if col["type"] == "number": payment_details[col["key"]] = _safe_float(val)
                        else: payment_details[col["key"]] = str(val).strip() if val is not None else None
                invoice_payload["payment_details"] = payment_details

                dispatch_details = {}
                for col in SALES_VOUCHER_COLUMNS:
                    if col["tab"] == "Dispatch Details":
                        val = first.get(col["label"])
                        if col["key"] == "dispatch_time": dispatch_details[col["key"]] = _get_time(val)
                        elif col["type"] == "date": dispatch_details[col["key"]] = _get_date(val)
                        elif col["type"] == "number": dispatch_details[col["key"]] = _safe_float(val)
                        else: dispatch_details[col["key"]] = str(val).strip() if val is not None else None
                
                rail_lr = str(first.get("Rail Upto Port - LR / GR / Consignment No") or "").strip()
                dispatch_details["rail_upto_port_lr_gr_consignment"] = rail_lr if rail_lr else None
                invoice_payload["dispatch_details"] = dispatch_details

                eway_details = []
                for s in range(1, 4):
                    prefix = f"EWB {s} - "
                    ewb_data_found = False
                    ewb_obj = { "eway_bill_available": True }
                    mapping = {
                        "E-Way Bill No": "eway_bill_no", "Date": "eway_bill_date", "Validity Period": "validity_period",
                        "Distance (km)": "distance", "Extension Date": "extension_date", "Extended EWB No": "extended_ewb_no",
                        "Extension Reason": "extension_reason", "From Place": "from_place", "Remaining Distance": "remaining_distance",
                        "New Validity": "new_validity", "Updated Vehicle No": "updated_vehicle_no"
                    }
                    for label_suffix, key in mapping.items():
                        val = first.get(prefix + label_suffix)
                        if val not in (None, ""):
                            ewb_data_found = True
                            ewb_obj[key] = _get_date(val) if "Date" in label_suffix else str(val).strip() # type: ignore
                        else: ewb_obj[key] = None # type: ignore
                    if ewb_data_found: eway_details.append(ewb_obj)
                
                print(f"DEBUG BRAIN: Voucher {v_no} -> extracted {len(eway_details)} E-Way Bills")
                invoice_payload["eway_bill_details"] = eway_details
                invoice_payload["irn"] = str(first.get("IRN") or "").strip()
                invoice_payload["ack_no"] = str(first.get("Acknowledgement No") or "").strip()

                # 6. Customer & Duplicate Invoice Validation
                val_res = validate_sales_customer_and_invoice(
                    tenant_id=tenant_id,
                    customer_name=invoice_payload.get("customer_name"),
                    gstin=invoice_payload.get("gstin"),
                    branch=invoice_payload.get("customer_branch"),
                    sales_invoice_no=v_no
                )

                if val_res["status"] != "READY":
                    creation_errors.append({
                        "row_num": first.get("_row_num"),
                        "voucher_number": v_no,
                        "customer": invoice_payload.get("customer_name", "N/A"),
                        "row_data": first, # Keep original row for report
                        "errors": {
                            "validation_error": [val_res["message"]]
                        },
                        "status": val_res["status"],
                        "matched_by": val_res.get("matched_by")
                    })
                    continue

                # Attach detected customer_id and matched branch name
                invoice_payload["customer_id"] = val_res.get("customer_id")
                if val_res.get("branch"):
                    invoice_payload["customer_branch"] = val_res.get("branch")

                # 7. Save Voucher (Single transaction per invoice)
                try:
                    with transaction.atomic():
                        serializer = VoucherSalesInvoiceDetailsSerializer(
                            data=invoice_payload, 
                            context={"request": request, "tenant_id": tenant_id}
                        )
                        if serializer.is_valid():
                            serializer.save()
                            actual_items = len(items) + len(foreign_items)
                            total_items_count += actual_items
                            created_vouchers.append({
                                "row_num": first.get("_row_num"),
                                "voucher_number": v_no,
                                "customer":       invoice_payload.get("customer_name", "N/A"),
                                "items_count":    actual_items,
                                "status": "Success"
                            })
                        else:
                            creation_errors.append({
                                "row_num": first.get("_row_num"),
                                "voucher_number": v_no,
                                "row_data": first,
                                "errors": serializer.errors,
                                "status": "Failed"
                            })
                except Exception as exc:
                    logger.error(f"Error saving voucher {v_no}: {exc}", exc_info=True)
                    creation_errors.append({
                        "row_num": first.get("_row_num"),
                        "voucher_number": v_no,
                        "row_data": first,
                        "errors": str(exc),
                        "status": "Failed"
                    })

            expected_labels_norm = {str(l).strip().upper() for l in HEADER_LABELS}
            extra_labels = list(uploaded_labels - expected_labels_norm)
            warnings = [f"Unknown columns ignored: {extra_labels}"] if extra_labels else []

            return Response({
                "message": "Excel processing complete.",
                "summary": {
                    "total_invoices_created": len(created_vouchers),
                    "total_items_inserted": total_items_count,
                    "failed_rows_count": len(creation_errors),
                },
                "created_vouchers": created_vouchers,
                "creation_errors": creation_errors,
                "warnings": warnings
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Excel upload failed: {e}", exc_info=True)
            return Response({"error": f"Internal error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ── Workflow Helpers ────────────────────────────────────────────────────────

def _safe_float(v):
    if v is None or v == "": return 0.0
    try: 
        if isinstance(v, str):
            v = v.replace(",", "").strip()
        return float(v)
    except: return 0.0

def _get_date(v, default=None):
    if v is None or v == "": return default
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str) and v.strip():
        try:
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
                try: return datetime.datetime.strptime(v.strip(), fmt).date()
                except: continue
        except: pass
    if isinstance(v, (int, float)):
        try:
            return (datetime.date(1899, 12, 30) + datetime.timedelta(days=v))
        except: return default
    return default # return None by default if not specified

def _get_time(v):
    if v is None or v == "": return None
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime.datetime):
        return v.time().strftime("%H:%M:%S")
    if isinstance(v, (int, float)):
        try:
            total_seconds = round(v * 24 * 3600)
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return datetime.time(hours % 24, minutes, seconds).strftime("%H:%M:%S")
        except: return None
    if isinstance(v, str) and v.strip():
        v_str = v.strip()
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
            try:
                clean_v = v_str.replace(".", ":") if "." in v_str and ":" not in v_str else v_str
                return datetime.datetime.strptime(clean_v, fmt).time().strftime("%H:%M:%S")
            except: continue
    return None

def process_sales_invoice_group(group_rows, tenant_id, session_id, row_index):
    """
    Groups line items into a single invoice structure and runs validation.
    """
    first = group_rows[0]
    invoice_no = str(first.get("Sales Invoice No.") or "").strip()
    
    # Header Mapping
    header = {}
    for col in SALES_VOUCHER_COLUMNS:
        # Columns that belong to Header (everything except item/tax/foreign currency)
        if col["tab"] not in ("Item & Tax Details", "Foreign Currency (Item & Tax Details)"):
            val = first.get(col["label"])
            key = col["key"]
            if col["type"] == "date":
                header[key] = str(_get_date(val)) if val else None
            elif col["type"] == "number":
                header[key] = _safe_float(val)
            else:
                header[key] = str(val).strip() if val else None

    # Normalization (as in regular upload)
    if not header.get("reverse_charge"): header["reverse_charge"] = "N"
    if not header.get("invoice_type"): header["invoice_type"] = "Regular"
    if not header.get("exchange_rate"): header["exchange_rate"] = 1.0

    # Line Items Extraction
    items = []
    for rd in group_rows:
        item = {}
        has_item = False
        for col in SALES_VOUCHER_COLUMNS:
            # Columns that belong to Line Items
            if col["tab"] in ("Item & Tax Details", "Foreign Currency (Item & Tax Details)"):
                val = rd.get(col["label"])
                if val not in (None, ""):
                    has_item = True
                    item[col["key"]] = _safe_float(val) if col["type"] == "number" else str(val).strip()
                else:
                    item[col["key"]] = 0.0 if col["type"] == "number" else None
        if has_item:
            items.append(item)

    # Validation
    val_res = validate_sales_customer_and_invoice(
        tenant_id=tenant_id,
        customer_name=header.get("customer_name"),
        gstin=header.get("gstin"),
        branch=header.get("customer_branch"),
        sales_invoice_no=invoice_no
    )

    # Required Fields Check — separate customer name from other required fields
    missing_core = []  # non-customer fields
    missing_customer_name = not header.get("customer_name")

    if not invoice_no: missing_core.append("Invoice No")
    if not header.get("gstin"): missing_core.append("GSTIN")
    if not header.get("customer_branch"): missing_core.append("Branch")

    status_val = val_res["status"]
    if missing_core:
        # Other critical fields missing → hard validation failure
        all_missing = (["Customer Name"] if missing_customer_name else []) + missing_core
        status_val = "VALIDATION_FAILED"
        val_res["message"] = f"Missing required fields: {', '.join(all_missing)}"
    elif missing_customer_name:
        # Only customer name is missing → treat as CUSTOMER_MISSING so user can create one
        status_val = "CUSTOMER_MISSING"
        val_res["message"] = "Customer name is missing. Please create a new customer."

    return {
        "invoice_no": invoice_no,
        "header": header,
        "items": items,
        "status": status_val,
        "message": val_res.get("message", ""),
        "customer_id": val_res.get("customer_id"),
        "matched_by": val_res.get("matched_by"),
        "session_id": session_id,
        "row_index": row_index
    }

class SalesExcelWorkflowUploadView(APIView):
    """
    Workflow Step 1: Upload Excel and return in-memory preview.
    Uses Django cache for temporary storage.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        tenant_id = getattr(request.user, "tenant_id", None)
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"error": "No file uploaded"}, status=400)

        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Map Excel Columns
            if ws is None:
                return Response({"error": "No active sheet found"}, status=400)
            
            excel_col_index = {str(cell.value).strip().upper(): idx for idx, cell in enumerate(ws[1]) if cell.value}
            header_map = {str(c["label"]): excel_col_index[str(c["label"]).upper()] 
                         for c in SALES_VOUCHER_COLUMNS if str(c["label"]).upper() in excel_col_index}

            # Batch Extract
            rows_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                row_dict = {label: row[idx] for label, idx in header_map.items()}
                rows_data.append(row_dict)

            # Group by Invoice No
            groups = {}
            last_no = None
            for rd in rows_data:
                curr_no = str(rd.get("Sales Invoice No.") or "").strip()
                if curr_no: last_no = curr_no
                if not last_no: continue
                if last_no not in groups: groups[last_no] = []
                groups[last_no].append(rd)

            session_id = str(uuid.uuid4())
            results = []
            for idx, (inv_no, g_rows) in enumerate(groups.items()):
                results.append(process_sales_invoice_group(g_rows, tenant_id, session_id, idx))

            cache.set(f"sales_upload_{session_id}", results, timeout=3600)
            return Response({"session_id": session_id, "invoices": results})

        except Exception as e:
            logger.error(f"Workflow upload error: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class SalesExcelWorkflowUpdateView(APIView):
    """
    Update / Revalidate an in-memory invoice.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            session_id = request.data.get("session_id")
            tenant_id = getattr(request.user, "tenant_id", None)
            
            # Scenario 1: Update Specific Invoice Data
            invoice_index = request.data.get("index") # List index in cache
            updated_data = request.data.get("invoice") # The whole invoice object from frontend

            # Scenario 2: Global Revalidation (e.g. after customer creation)
            revalidate_all = request.data.get("revalidate_all", False)

            cache_key = f"sales_upload_{session_id}"
            invoices = cache.get(cache_key)
            if not invoices:
                return Response({"error": "Session expired"}, status=404)

            if revalidate_all:
                # Re-run validation for all
                for inv in invoices:
                    # Re-validate using stored header
                    header = inv.get("header", {})
                    val_res = validate_sales_customer_and_invoice(
                        tenant_id=tenant_id,
                        customer_name=str(header.get("customer_name") or ""),
                        gstin=str(header.get("gstin") or ""),
                        branch=str(header.get("customer_branch") or ""),
                        sales_invoice_no=str(inv.get("invoice_no") or "")
                    )
                    inv["status"] = val_res["status"]
                    inv["message"] = val_res.get("message", "")
                    inv["customer_id"] = val_res.get("customer_id")
            elif invoice_index is not None and updated_data:
                # Update specific record and revalidate
                header = updated_data.get("header", {})
                val_res = validate_sales_customer_and_invoice(
                    tenant_id=tenant_id,
                    customer_name=str(header.get("customer_name") or ""),
                    gstin=str(header.get("gstin") or ""),
                    branch=str(header.get("customer_branch") or ""),
                    sales_invoice_no=str(updated_data.get("invoice_no") or "")
                )
                updated_data["status"] = val_res["status"]
                updated_data["message"] = val_res.get("message", "")
                updated_data["customer_id"] = val_res.get("customer_id")
                invoices[invoice_index] = updated_data

            cache.set(cache_key, invoices, timeout=3600)
            return Response({"invoices": invoices})
        except Exception as e:
            logger.error(f"Sales workflow update error: {e}", exc_info=True)
            return Response({"error": str(e)}, status=400)

class SalesExcelWorkflowFinalizeView(APIView):
    """
    Finalize ONLY 'READY' invoices.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            session_id = request.data.get("session_id")
            tenant_id = getattr(request.user, "tenant_id", 1)
            
            cache_key = f"sales_upload_{session_id}"
            invoices = cache.get(cache_key)
            if not invoices:
                return Response({"error": "Session expired"}, status=404)

            summary = {"total": len(invoices), "created": 0, "failed": 0, "skipped": 0, "errors": []}

            for inv in invoices:
                if inv["status"] != "READY":
                    summary["skipped"] += 1
                    continue

                try:
                    with transaction.atomic():
                        # Map workflow structure to Sales Voucher Serializer
                        header = inv["header"]
                        items = inv["items"]
                        
                        payload = {}
                        payment_details = {}
                        dispatch_details = {}
                        eway_bill_details = []
                        
                        # Group EWB entries
                        ewb_entries = {1: {}, 2: {}, 3: {}}

                        # Valid keys for the main Invoice Details model from serializer
                        valid_header_keys = [
                            'date', 'sales_invoice_no', 'voucher_name', 'outward_slip_no',
                            'customer_name', 'customer_id', 'customer_branch', 'bill_to', 'ship_to', 
                            'gstin', 'contact', 'tax_type', 'state_type', 'export_type', 
                            'exchange_rate', 'sales_order_no', 'place_of_supply', 'reverse_charge', 
                            'invoice_type', 'gst_export_type', 'port_code', 'shipping_bill_number', 
                            'shipping_bill_date', 'ecommerce_gstin', 'irn', 'ack_no'
                        ]

                        for col in SALES_VOUCHER_COLUMNS:
                            key = col["key"]
                            val = header.get(key)
                            if val in (None, ""): continue

                            # Pre-process dates
                            if col["type"] == "date":
                                val = str(_get_date(val))

                            if col["tab"] == "Payment Details":
                                payment_details[key] = val
                            elif col["tab"] == "Dispatch Details":
                                dispatch_details[key] = val
                            elif col["tab"] == "E-Invoice & E-Way Bill Details":
                                k_str = str(key)
                                if k_str.startswith("ewb1_"):
                                    ewb_entries[1][k_str.replace("ewb1_", "")] = val
                                elif k_str.startswith("ewb2_"):
                                    ewb_entries[2][k_str.replace("ewb2_", "")] = val
                                elif k_str.startswith("ewb3_"):
                                    ewb_entries[3][k_str.replace("ewb3_", "")] = val
                                elif key in valid_header_keys:
                                    payload[key] = val
                            elif key in valid_header_keys:
                                if key == "date":
                                    payload[key] = val
                                else:
                                    payload[key] = val

                        # Finalize EWB
                        for e_idx in (1, 2, 3):
                            if ewb_entries[e_idx]:
                                # Map "available" string to boolean
                                avail = ewb_entries[e_idx].get("available", "N")
                                # Map keys to model fields
                                mapped_ewb = {
                                    "eway_bill_available": (str(avail).upper() in ("Y", "YES", "TRUE", "1")),
                                    "eway_bill_no": ewb_entries[e_idx].get("eway_bill_no"),
                                    "eway_bill_date": _get_date(ewb_entries[e_idx].get("date")),
                                    "validity_period": ewb_entries[e_idx].get("validity_period"),
                                    "distance": ewb_entries[e_idx].get("distance"),
                                    "extension_date": _get_date(ewb_entries[e_idx].get("extension_date")),
                                    "extended_ewb_no": ewb_entries[e_idx].get("extended_ewb_no"),
                                    "extension_reason": ewb_entries[e_idx].get("extension_reason"),
                                    "from_place": ewb_entries[e_idx].get("from_place"),
                                    "remaining_distance": ewb_entries[e_idx].get("remaining_distance"),
                                    "new_validity": ewb_entries[e_idx].get("new_validity"),
                                    "updated_vehicle_no": ewb_entries[e_idx].get("updated_vehicle_no"),
                                }
                                # Only add if has at least an EWB number or is explicitly available
                                if mapped_ewb["eway_bill_no"] or mapped_ewb["eway_bill_available"]:
                                    eway_bill_details.append(mapped_ewb)

                        # Concatenate Addresses
                        bill_to_parts = [header.get(k) for k in ["bill_to_address_1", "bill_to_address_2", "bill_to_address_3", "bill_to_city", "bill_to_state", "bill_to_pincode", "bill_to_country"] if header.get(k)]
                        if bill_to_parts: payload["bill_to"] = ", ".join(map(str, bill_to_parts))
                        
                        ship_to_parts = [header.get(k) for k in ["ship_to_address_1", "ship_to_address_2", "ship_to_address_3", "ship_to_city", "ship_to_state", "ship_to_pincode", "ship_to_country"] if header.get(k)]
                        if ship_to_parts: payload["ship_to"] = ", ".join(map(str, ship_to_parts))

                        # Process Items (Standard vs Foreign)
                        std_items = []
                        foreign_items = []
                        
                        for itm in items:
                            is_foreign = any(k.startswith("fc_") for k in itm.keys() if itm.get(k))
                            if is_foreign:
                                mapped_fc = {
                                    "item_name": itm.get("fc_item_name"),
                                    "description": itm.get("fc_description"),
                                    "quantity": itm.get("fc_qty"),
                                    "uqc": itm.get("fc_uom"),
                                    "rate": itm.get("fc_item_rate"),
                                    "amount": itm.get("fc_invoice_value"),
                                    "alternate_unit": itm.get("fc_alternate_unit"),
                                    "sales_ledger": itm.get("fc_sales_ledger"),
                                }
                                if mapped_fc["item_name"]: foreign_items.append(mapped_fc)
                            else:
                                if itm.get("item_name"): std_items.append(itm)

                        payload["items"] = std_items
                        payload["foreign_items"] = foreign_items
                        payload["customer_id"] = inv["customer_id"]
                        payload["payment_details"] = payment_details
                        payload["dispatch_details"] = dispatch_details
                        payload["eway_bill_details"] = eway_bill_details

                        serializer = VoucherSalesInvoiceDetailsSerializer(
                            data=payload, 
                            context={"request": request, "tenant_id": tenant_id}
                        )
                        if serializer.is_valid():
                            serializer.save()
                            summary["created"] += 1
                            inv["created"] = True
                        else:
                            summary["failed"] += 1
                            summary["errors"].append({
                                "invoice_no": inv["invoice_no"],
                                "errors": serializer.errors
                            })
                            inv["status"] = "VALIDATION_FAILED"
                            # Store field-level errors for frontend highlighting
                            inv["field_errors"] = serializer.errors
                            inv["message"] = "Voucher creation failed. Click Edit & Fix to resolve errors."
                except Exception as e:
                    logger.error(f"Finalize error for {inv.get('invoice_no')}: {e}", exc_info=True)
                    summary["failed"] += 1
                    summary["errors"].append({"invoice_no": inv["invoice_no"], "errors": str(e)})
                    inv["status"] = "VALIDATION_FAILED"
                    inv["message"] = f"System Error: {str(e)}"

            # Remove from cache if all processed? Or partial?
            remaining = [inv for inv in invoices if not inv.get("created")]
            if not remaining:
                cache.delete(cache_key)
            else:
                # Update statuses for those that failed? 
                # (Actually inv["status"] didn't change unless it failed inside loop)
                cache.set(cache_key, remaining, timeout=3600)

            return Response({"summary": summary, "remaining": remaining})
        except Exception as e:
            logger.error(f"Sales workflow finalize global error: {e}", exc_info=True)
            return Response({"error": str(e)}, status=400)
